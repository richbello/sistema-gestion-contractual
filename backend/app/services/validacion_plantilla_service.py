# -*- coding: utf-8 -*-
"""
Validacion de plantilla de pagos (V1): caracteres/SAP + estructura.
No modifica la plantilla; devuelve un informe estructurado.
"""
import re
import io
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .normalizacion import normalizar_nombre, quitar_tildes

INVISIBLES = {
    "espacio no-separable (NBSP)": "\u00a0",
    "espacio fino": "\u2009",
    "espacio de ancho cero": "\u200b",
    "BOM / ZWNBSP": "\ufeff",
    "salto de linea (LF)": "\n",
    "retorno de carro (CR)": "\r",
    "tabulacion": "\t",
}
TIPOGRAFICAS = {
    "\u201c": '"', "\u201d": '"',
    "\u2018": "'", "\u2019": "'",
    "\u2013": "-", "\u2014": "-",
    "\u2026": "...", "\u00b4": "'",
}
SIMBOLOS_RIESGO = set("&#%*;|\\")
MOJIBAKE = re.compile(r"[\u00c3\u00c2][\x80-\xbf\u00a1-\u00bf]")


def _analizar_texto(valor):
    problemas = []
    for nombre, ch in INVISIBLES.items():
        if ch in valor:
            problemas.append(nombre)
    if any(c in TIPOGRAFICAS for c in valor):
        problemas.append("caracter tipografico de Office")
    riesgo = sorted({c for c in valor if c in SIMBOLOS_RIESGO})
    if riesgo:
        problemas.append("simbolo de riesgo SAP: " + " ".join(riesgo))
    if MOJIBAKE.search(valor):
        problemas.append("posible mojibake (doble codificacion)")
    if valor != valor.strip():
        problemas.append("espacios al inicio/final")
    if "  " in valor:
        problemas.append("espacios dobles")
    for c in valor:
        if unicodedata.category(c).startswith("C") and c not in INVISIBLES.values():
            problemas.append(f"caracter de control U+{ord(c):04X}")
            break
    acentuados = sorted({c for c in valor
                         if unicodedata.combining(unicodedata.normalize("NFKD", c)[-1])})
    if acentuados:
        problemas.append("tilde o enie que SAP rechaza: " + " ".join(acentuados))
    return problemas


def _num(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def _solo_digitos(v):
    return re.sub(r"\D", "", str(v or ""))


def _generar_saneada(ruta_plantilla):
    """Crea una copia del .xlsx con tildes/enie fuera en todas las columnas
    de texto y normalizar_nombre completo solo en la columna K (nombre).
    Devuelve (bytes_del_archivo, lista_de_cambios)."""
    wb = load_workbook(ruta_plantilla, data_only=True)
    ws = wb.active
    cambios = []
    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            if not isinstance(celda.value, str):
                continue
            original = celda.value
            if celda.column == 11:  # columna K = nombre del beneficiario
                nuevo = normalizar_nombre(original)
            else:
                nuevo = quitar_tildes(original)
            if nuevo != original:
                ref = f"{get_column_letter(celda.column)}{celda.row}"
                cambios.append({"celda": ref, "antes": original[:60], "despues": nuevo[:60]})
                celda.value = nuevo
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), cambios

def validar_plantilla(ruta_plantilla):
    wb = load_workbook(ruta_plantilla, data_only=True)
    ws = wb.active

    caracteres = []
    estructura = []
    filas = list(ws.iter_rows(min_row=2))

    for fila in filas:
        for celda in fila:
            if isinstance(celda.value, str):
                probs = _analizar_texto(celda.value)
                if probs:
                    ref = f"{get_column_letter(celda.column)}{celda.row}"
                    caracteres.append({
                        "celda": ref,
                        "valor": celda.value[:60],
                        "problemas": probs,
                    })

    bloques_ok = 0
    bloques_total = 0
    i = 0
    n = len(filas)
    while i < n:
        r = filas[i]
        tipo = str(r[0].value or "").strip().upper()
        if tipo == "C":
            bloques_total += 1
            fila_excel = r[0].row
            r40 = filas[i + 1] if i + 1 < n else None
            r31 = filas[i + 2] if i + 2 < n else None
            clave40 = str(r40[1].value).strip() if r40 and r40[1].value is not None else ""
            clave31 = str(r31[1].value).strip() if r31 and r31[1].value is not None else ""

            problemas_bloque = []
            if clave40 != "40":
                problemas_bloque.append("falta o descuadra la fila P40")
            if clave31 != "31":
                problemas_bloque.append("falta o descuadra la fila P31")

            if not problemas_bloque:
                ced = _solo_digitos(r31[4].value)
                if not ced:
                    problemas_bloque.append("cedula del P31 vacia")
                elif len(ced) < 6 or len(ced) > 11:
                    problemas_bloque.append(f"cedula con longitud sospechosa ({len(ced)}): {ced}")
                banco = _solo_digitos(r31[36].value)
                cuenta = _solo_digitos(r31[37].value)
                tipo_cta = _solo_digitos(r31[38].value)
                if not banco:
                    problemas_bloque.append("codigo de banco vacio")
                if not cuenta:
                    problemas_bloque.append("numero de cuenta vacio")
                if not tipo_cta:
                    problemas_bloque.append("tipo de cuenta vacio")
                imp40 = _num(r40[7].value)
                imp31 = _num(r31[7].value)
                if imp40 is None or imp31 is None:
                    problemas_bloque.append("importe vacio en P40 o P31")
                elif abs(imp40 - imp31) > 0.5:
                    problemas_bloque.append(f"importe P40 ({imp40:,.0f}) != P31 ({imp31:,.0f})")

            if problemas_bloque:
                nombre = str(r[10].value or "").strip()[:40]
                estructura.append({
                    "fila": fila_excel,
                    "contratista": nombre,
                    "problemas": problemas_bloque,
                })
            else:
                bloques_ok += 1
            i += 3
        else:
            i += 1

    import base64
    contenido, cambios = _generar_saneada(ruta_plantilla)
    plantilla_b64 = base64.b64encode(contenido).decode("ascii")

    return {
        "ok": True,
        "archivo": ruta_plantilla,
        "resumen": {
            "bloques_total": bloques_total,
            "bloques_ok": bloques_ok,
            "bloques_con_error": bloques_total - bloques_ok,
            "celdas_con_caracteres": len(caracteres),
        },
        "estructura": estructura,
        "caracteres": caracteres[:200],
        "caracteres_truncado": len(caracteres) > 200,
        "cambios_saneado": cambios,
        "total_cambios": len(cambios),
        "plantilla_saneada_b64": plantilla_b64,
    }