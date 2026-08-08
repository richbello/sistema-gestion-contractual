# -*- coding: utf-8 -*-
"""
Módulo 12 · CDP-CRP de OXP  ─  Servicio backend (CRP + CDP)
====================================================================
Rellena las hojas 'CRP_vf' y 'CDP' de la plantilla a partir de filas ya
FILTRADAS (saldo > 0) y MAPEADAS en el navegador (claves canónicas). El backend
agrupa por CRP, aplica constantes/conversiones y escribe → RAM mínima (Render).

AGRUPAMIENTO (según formato diligenciado):
    Un CRP puede tener varias POSICIONES. Se agrupa por 'interno_crp'
    (N° Interno CRP del reporte); cada grupo recibe UN número consecutivo
    (1..N) y sus filas conservan la 'pos_crp' (N° Posición CRP).

CDP:
    - Posición Presupuestal = Nuevo Rubro (conversión RUBRO_MAPEO).
    - Elemento PEP = según rubro convertido (ELEMENTO_PEP_POR_RUBRO).
    - Fondos = 1-200-I071 · Cuenta de Mayor = 7990990000 (fijos).
    - Objeto = "REEMPLAZA CDP N Y CRP N AL CONSTITUIRSE ... CUYO OBJETO ES <obj>"
      (en OXP el nuevo CDP y el nuevo CRP comparten el mismo consecutivo N).
    - Fecha Oficio = hoy · Num. Ext. Entidad = consecutivo por fila desde 1.

Contrato de datos (cada fila = dict con claves canónicas; opcionales):
    importe, objeto, tipo_compromiso, no_compromiso, modo_seleccion,
    tipo_doc_benef, id_benef, id_solicitante, id_responsable,
    interno_crp, pos_crp, num_crp, rubro, elemento_pep, fondos
"""

import os
from io import BytesIO
from datetime import date

import openpyxl

# ----------------------------------------------------------------------------
# Constantes (revisar por alcaldía)
# ----------------------------------------------------------------------------
PLANTILLA_PATH = os.environ.get(
    "OXP_PLANTILLA_PATH",
    os.path.join(os.path.dirname(__file__), "..", "plantillas",
    "Formato_CRP_de_OXP.xlsx"),
)

SOCIEDAD_FIJA   = 1001            # Usme = 1001 (confirmado)
CLASE_DOC_CRP   = "RP"
MONEDA          = "COP"
POSICION_CDP    = 1
TIPO_DE_PAGO    = "02"
FECHA_INICIAL   = "01/01/2026"
FECHA_FINAL     = "31/12/2026"

# Mapeo de Rubros: Rubro del reporte CRP -> Nuevo Rubro para CDP (Posicion Presupuestal)
RUBRO_MAPEO = {
    "O2110103007": "O219001",
    "O2120201003023212901": "O219001",
    "O2120201003023215202": "O219001",
    "O2120201003023215307": "O219001",
    "O2120201003023262005": "O219001",
    "O2120201003023270101": "O219001",
    "O2120201003053511007": "O219001",
    "O2120201003053511036": "O219001",
    "O2120201003053529901": "O219001",
    "O2120201003053542006": "O219001",
    "O2120201003053542007": "O219001",
    "O2120201003063627018": "O219001",
    "O2120201003063692002": "O219001",
    "O2120201003063699005": "O219001",
    "O2120201003063699006": "O219001",
    "O2120201003063699046": "O219001",
    "O2120201003063699060": "O219001",
    "O2120201003083811106": "O219001",
    "O2120201003083891102": "O219001",
    "O2120201003083891104": "O219001",
    "O2120201003083891106": "O219001",
    "O2120201004024291304": "O219001",
    "O2120201004024291305": "O219001",
    "O2120201004024291501": "O219001",
    "O2120201004024292202": "O219001",
    "O2120201004024299206": "O219001",
    "O2120201004024299209": "O219001",
    "O2120201004024299214": "O219001",
    "O2120201004024299917": "O219001",
    "O2120201004054516003": "O219001",
    "O2120201004054516004": "O219001",
    "O2120201004054516005": "O219001",
    "O2120201004064613201": "O219001",
    "O2120201004064653102": "O219001",
    "O21202020060363399": "O219001",
    "O21202020060868021": "O219001",
    "O212020200662165": "O219001",
    "O21202020070103010271311": "O219001",
    "O212020200701030471349": "O219001",
    "O212020200701030571354": "O219001",
    "O212020200801787143": "O219001",
    "O21202020080383143": "O219001",
    "O21202020080484120": "O219001",
    "O21202020080585250": "O219001",
    "O21202020080585330": "O219001",
    "O21202020080686312": "O219001",
    "O21202020080686320": "O219001",
    "O21202020080686330": "O219001",
    "O21202020080787130": "O219001",
    "O2120202008078714102": "O219001",
    "O2120202008078715202": "O219001",
    "O2120202008078715701": "O219001",
    "O21202020090494110": "O219001",
    "O21202020090494239": "O219001",
    "O219001": "O219002",
    "O219002": "O219002",
    "O23011745992024236401000": "O230689",
    "O23011745992024238401000": "O230689",
    "O23011745992024239601000": "O230689",
    "O23011745992024241301000": "O230689",
    "O23011745992024241801000": "O230689",
    "O23011745992024242201000": "O230689",
    "O23011745992024242601000": "O230689",
    "O23011745992024242901000": "O230689",
    "O23011745992024243401000": "O230689",
    "O23011745992024243801000": "O230689",
    "O23011745992024249301000": "O230689",
    "O23011745992024251401000": "O230689",
    "O23011745992024253101000": "O230689",
    "O23011745992024254901000": "O230689",
    "O23011745992024256201000": "O230689",
    "O23011745992024257001000": "O230689",
    "O23011745992024257101000": "O230689",
    "O23011745992024268501000": "O230689",
    "O23011745992024269101000": "O230689",
    "O23011745992024269201000": "O230689",
    "O23011745992024269801000": "O230689",
    "O23011745992024269901000": "O230689",
    "O23011745992024273101000": "O230689",
    "O23011745992024280601000": "O230689",
    "O23011745992024282101000": "O230689",
    "O23011745992024282201000": "O230689",
    "O23011745992024282301000": "O230689",
    "O23011745992024293201000": "O230689",
    "O230689": "O230690",
    "O230690": "O230690",
}

# Mapeo de Elemento PEP por Rubro convertido
# Basado en la tabla proporcionada: cada Posición Presupuestal tiene su Elemento PEP asociado
ELEMENTO_PEP_POR_RUBRO = {
    "O219002": "PO/0005/0001/0000000005",
    "O230690": "PO/0005/0001/OBLI_INV",
    "O219001": "PO/0005/0001/0000000005",
    "O230689": "PO/0005/0001/OBLI_INV_VI",
}

def _convertir_rubro(rubro_origen):
    """Convierte rubro del reporte CRP a nuevo rubro para CDP."""
    if not rubro_origen:
        return None
    v = str(rubro_origen).strip()
    return RUBRO_MAPEO.get(v, v)   # si no está en el mapeo, deja el original


def _extraer_tipo_contrato(objeto):
    """Extrae el tipo de contrato del objeto original.

    En el original aparece '... POR PAGAR EL/LA <TIPO> <No>-<año> CUYO OBJETO ES ...'.
    Se toma <TIPO> (p.ej. 'CONTRATO DE OBRA', 'CONTRATO DE PRESTACION DE SERVICIOS'):
    las palabras tras 'POR PAGAR EL/LA' hasta el primer token con dígito o 'CUYO'.
    """
    if not objeto:
        return "CONTRATO DE PRESTACION DE SERVICIOS"
    txt = str(objeto)
    marcador = "POR PAGAR EL/LA"
    idx = txt.find(marcador)
    if idx == -1:
        return "CONTRATO DE PRESTACION DE SERVICIOS"
    resto = txt[idx + len(marcador):].strip()
    palabras = []
    for w in resto.split():
        wu = w.upper()
        if wu.startswith("CUYO"):
            break
        if any(ch.isdigit() for ch in w):
            break
        palabras.append(w)
    tipo = " ".join(palabras).strip()
    return tipo or "CONTRATO DE PRESTACION DE SERVICIOS"


def _reconstruir_objeto(num_cdp, num_crp, objeto_original):
    """Reconstruye la columna Objeto con formato OXP.

    Patrón (verificado contra referencia):
    REEMPLAZA CDP <num_cdp> Y CRP <num_crp> AL CONSTITUIRSE COMO OBLIGACION
    POR PAGAR EL/LA <TIPO_CONTRATO> CUYO OBJETO ES <objeto_original>

    <TIPO_CONTRATO> se extrae del propio objeto original.
    """
    if not objeto_original:
        return ""

    num_cdp_str = str(num_cdp).strip() if num_cdp not in (None, "") else ""
    num_crp_str = str(num_crp).strip() if num_crp not in (None, "") else ""
    obj_str = str(objeto_original).strip()

    if not num_cdp_str or not num_crp_str:
        return obj_str

    tipo = _extraer_tipo_contrato(obj_str)
    return (f"REEMPLAZA CDP {num_cdp_str} Y CRP {num_crp_str} AL CONSTITUIRSE COMO "
            f"OBLIGACION POR PAGAR EL/LA {tipo} CUYO OBJETO ES {obj_str}")


def _obtener_elemento_pep(rubro_convertido):
    """Obtiene el Elemento PEP correspondiente al rubro convertido.
    
    Si no está en la tabla, retorna None (campo en blanco).
    """
    if not rubro_convertido:
        return None
    v = str(rubro_convertido).strip()
    return ELEMENTO_PEP_POR_RUBRO.get(v)


# Num. Ext. Entidad (columna V). Modos:
#   "num_crp" -> número de CRP original del reporte (constante por grupo)  [segun imagen]
#   "grupo"   -> mismo consecutivo del CRP (1..N por grupo)
#   "fila"    -> consecutivo por fila (1..N)   [texto original del spec]
NUM_EXT_ENTIDAD_MODO = "num_crp"


def _hoy():
    return date.today().strftime("%d.%m.%Y")


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str):
        limpio = v.replace(".", "").replace(",", "").strip()
        try:
            v = float(limpio)
        except ValueError:
            return v
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return v


def _g(fila, clave):
    v = fila.get(clave)
    return None if v in ("", None) else v


def _pos_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 999999   # posiciones sin dato van al final del grupo


def _agrupar_por_crp(filas):
    """Agrupa por interno_crp preservando el orden de primera aparicion.
    Cada grupo: filas ordenadas por pos_crp. Devuelve lista de grupos."""
    grupos = {}
    orden = []
    for i, fila in enumerate(filas):
        clave = _g(fila, "interno_crp")
        if clave is None:
            clave = "__sin_interno_%d" % i   # fila suelta = su propio grupo
        if clave not in grupos:
            grupos[clave] = []
            orden.append(clave)
        grupos[clave].append(fila)
    return [sorted(grupos[k], key=lambda f: _pos_int(_g(f, "pos_crp"))) for k in orden]


# ----------------------------------------------------------------------------
# CRP OXP  -> hoja CRP_vf
# ----------------------------------------------------------------------------
def llenar_crp_vf(filas, plantilla_path=None):
    plantilla_path = plantilla_path or PLANTILLA_PATH
    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb["CRP_vf"]
    hoy = _hoy()

    grupos = _agrupar_por_crp(filas)

    r = 2
    for crp_consecutivo, grupo in enumerate(grupos, start=1):
        for pos_idx, fila in enumerate(grupo, start=1):
            pos = _g(fila, "pos_crp")
            posicion = _num(pos) if pos not in (None, "") else pos_idx

            if NUM_EXT_ENTIDAD_MODO == "num_crp":
                num_ext = _g(fila, "num_crp")
            elif NUM_EXT_ENTIDAD_MODO == "grupo":
                num_ext = crp_consecutivo
            else:
                num_ext = r - 1

            ws.cell(row=r, column=1,  value=crp_consecutivo)            # A CRP (grupo)
            ws.cell(row=r, column=2,  value=posicion)                   # B Posicion
            ws.cell(row=r, column=3,  value=hoy)                        # C Fecha Documento
            ws.cell(row=r, column=4,  value=hoy)                        # D Fecha Contab.
            ws.cell(row=r, column=5,  value=SOCIEDAD_FIJA)             # E Sociedad
            ws.cell(row=r, column=6,  value=CLASE_DOC_CRP)            # F Clase Documento
            ws.cell(row=r, column=7,  value=MONEDA)                    # G Moneda
            ws.cell(row=r, column=8,  value=_num(_g(fila, "importe")))  # H importe
            # I CDP -> en blanco (por ahora)
            ws.cell(row=r, column=10, value=POSICION_CDP)             # J Posicion CDP
            ws.cell(row=r, column=11, value=_g(fila, "objeto"))         # K Objeto
            ws.cell(row=r, column=12, value=_g(fila, "tipo_compromiso"))# L Tipo compromiso
            ws.cell(row=r, column=13, value=_g(fila, "no_compromiso"))  # M No. Compromiso
            ws.cell(row=r, column=14, value=FECHA_INICIAL)            # N Fecha Inicial
            ws.cell(row=r, column=15, value=FECHA_FINAL)              # O Fecha Final
            cp = ws.cell(row=r, column=16, value=TIPO_DE_PAGO)       # P Tipo de Pago
            cp.number_format = "@"
            ws.cell(row=r, column=17, value=_g(fila, "modo_seleccion"))  # Q Modo Seleccion
            ws.cell(row=r, column=18, value=_g(fila, "tipo_doc_benef"))  # R Tipo Doc Benef
            ws.cell(row=r, column=19, value=_g(fila, "id_benef"))        # S Identif. Benef
            ws.cell(row=r, column=20, value=_g(fila, "id_solicitante"))  # T ID Solicitante
            ws.cell(row=r, column=21, value=_g(fila, "id_responsable"))  # U ID Responsable
            ws.cell(row=r, column=22, value=num_ext)                    # V Num. Ext. Entidad
            r += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, (r - 2)


# ----------------------------------------------------------------------------
# CDP OXP  -> hoja CDP
# Estructura real de la plantilla (columnas A..S):
#   A CDP | B Posicion | C Fecha Documento | D Fecha Contabilizacion |
#   E Clase Documento | F Sociedad | G Moneda | H importe Original |
#   I Posicion Presupuestal | J Fondos | K Elemento PEP | L Periodo Presupuestal |
#   M Cuenta de Mayor | N Objeto | O Numero Oficio | P Fecha Oficio |
#   Q ID Solicitante | R ID Responsable | S Num. Ext. Entidad
#
# Reglas:
#   - CDP (A) = consecutivo por grupo (interno_crp). En una OXP el nuevo CDP y
#     el nuevo CRP comparten el MISMO numero, por eso el Objeto usa este mismo
#     consecutivo para "CDP N" y "CRP N".
#   - Posicion Presupuestal (I) = Nuevo Rubro (conversion RUBRO_MAPEO).
#   - Fondos (J) = fijo "1-200-I071".
#   - Elemento PEP (K) = segun rubro convertido (ELEMENTO_PEP_POR_RUBRO).
#   - Cuenta de Mayor (M) = fijo CUENTA_MAYOR_FIJA.
#   - Objeto (N) = "REEMPLAZA CDP N Y CRP N AL CONSTITUIRSE ... CUYO OBJETO ES <obj>".
#   - Numero Oficio (O) = fijo "constitución OxP".
#   - Fecha Oficio (P) = hoy.
#   - Num. Ext. Entidad (S) = consecutivo por fila desde 1.
# ----------------------------------------------------------------------------
FONDOS_FIJO       = "1-200-I071"
CUENTA_MAYOR_FIJA = 7990990000            # numérico (confirmado en referencia)
NUMERO_OFICIO     = "constitución OxP"


def _calcular_posicion_presupuestal(nuevo_rubro, rubro, elemento_pep):
    """Determina la Posición Presupuestal del CDP (regla verificada 886/886).

    1. Si 'Nuevo Rubro' (col R del reporte) viene lleno -> se usa tal cual.
    2. Si el Elemento PEP del reporte empieza con 'PM' (inversión) -> O230689.
    3. Si el Elemento PEP contiene 'OBLI_INV' -> O230690.
    4. Si el Elemento PEP es de funcionamiento (PO/.../0000000005):
         rubros cortos (O219001/O219002) -> O219002 ; detallados -> O219001
         (se resuelve con RUBRO_MAPEO).
    5. Fallback -> rubro original.
    """
    if nuevo_rubro not in (None, ""):
        return str(nuevo_rubro).strip()
    pep = str(elemento_pep).strip() if elemento_pep else ""
    if pep.startswith("PM"):
        return "O230689"
    if "OBLI_INV" in pep:
        return "O230690"
    if pep == "PO/0005/0001/0000000005":
        return RUBRO_MAPEO.get(str(rubro).strip(), "O219001")
    return str(rubro).strip() if rubro else None


def llenar_cdp_oxp(filas, plantilla_path=None):
    """Rellena la hoja CDP directamente desde las filas del reporte CRP.

    Cada fila (dict con claves canónicas) representa un registro OXP ya filtrado
    (Com.Sin.Aut.Giro > 0). No se agrupa: una fila del reporte = una fila del CDP,
    en el mismo orden. Reglas verificadas 886/886 contra el formato de referencia.
    """
    plantilla_path = plantilla_path or PLANTILLA_PATH
    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb["CDP"]

    r = 2               # fila de datos (encabezado en fila 1)
    num_ext = 1         # Num. Ext. Entidad: consecutivo global desde 1
    hoy = _hoy()                    # fecha REAL del día en que se genera el CDP
    periodo = date.today().year     # período = año en curso al generar
    for fila in filas:
        num_cdp = _g(fila, "num_cdp")
        num_crp = _g(fila, "num_crp")

        # Posición Presupuestal (I): regla verificada
        pos_pres = _calcular_posicion_presupuestal(
            _g(fila, "nuevo_rubro"), _g(fila, "rubro"), _g(fila, "elemento_pep"))

        # Elemento PEP (K): derivado del rubro final
        elemento_pep = _obtener_elemento_pep(pos_pres)

        # Objeto (N): "REEMPLAZA CDP <n> Y CRP <n> AL CONSTITUIRSE ... CUYO OBJETO ES <obj>"
        objeto_txt = _reconstruir_objeto(num_cdp, num_crp, _g(fila, "objeto"))

        ws.cell(row=r, column=1,  value=_num(num_cdp))              # A CDP (Número de CDP del reporte)
        ws.cell(row=r, column=2,  value=1)                         # B Posición (siempre 1)
        ws.cell(row=r, column=3,  value=hoy)                       # C Fecha Documento (hoy)
        ws.cell(row=r, column=4,  value=hoy)                       # D Fecha Contabilización (hoy)
        ws.cell(row=r, column=5,  value="CP")                      # E Clase Documento
        ws.cell(row=r, column=6,  value=SOCIEDAD_FIJA)             # F Sociedad
        ws.cell(row=r, column=7,  value=MONEDA)                    # G Moneda
        ws.cell(row=r, column=8,  value=_num(_g(fila, "importe")))  # H importe Original
        ws.cell(row=r, column=9,  value=pos_pres)                 # I Posición Presupuestal
        ws.cell(row=r, column=10, value=FONDOS_FIJO)              # J Fondos
        if elemento_pep:
            ws.cell(row=r, column=11, value=elemento_pep)         # K Elemento PEP
        ws.cell(row=r, column=12, value=periodo)                  # L Período Presupuestal (año actual)
        ws.cell(row=r, column=13, value=CUENTA_MAYOR_FIJA)        # M Cuenta de Mayor
        ws.cell(row=r, column=14, value=objeto_txt)               # N Objeto
        ws.cell(row=r, column=15, value=NUMERO_OFICIO)            # O Número Oficio
        ws.cell(row=r, column=16, value=hoy)                      # P Fecha Oficio (hoy)
        ws.cell(row=r, column=17, value=_g(fila, "id_solicitante"))  # Q ID Solicitante
        ws.cell(row=r, column=18, value=_g(fila, "id_responsable"))  # R ID Responsable
        ws.cell(row=r, column=19, value=num_ext)                  # S Num. Ext. Entidad

        r += 1
        num_ext += 1

    # Dejar solo la hoja CDP en el archivo entregado (sin Instrucciones/AjusteValor/CRP/CRP_vf)
    for nombre_hoja in list(wb.sheetnames):
        if nombre_hoja != "CDP":
            del wb[nombre_hoja]
    ws.sheet_state = "visible"
    wb.active = 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, (r - 2)
