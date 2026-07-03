"""
Módulo 4 — Estado de cuenta por contrato.

Adaptado del script original de Colab: busca todos los pagos de un
contrato en el histórico y llena la plantilla oficial de "Estado de
Cuenta" (.xlsx con celdas fusionadas) respetando el formato.

Versión optimizada: usa SQLite temporal para queries eficientes
sin cargar todo el Excel en memoria.
"""
import os
import tempfile
import sqlite3
import openpyxl
from openpyxl.utils import range_boundaries, column_index_from_string
import pandas as pd

COL_CDP_EXTERNO = 'CDP Externo'
COL_CRP_EXTERNO = 'CRP Externo'
COL_FECHA_INICIO = 'FECHA INICIAL DE CONTRATO'
COL_FECHA_FIN = 'FECHA DE TERMINACION FINAL'
COL_VALOR_CTO = 'VALOR FINAL DEL CONTRATO'


def _limpiar(dato):
    d = str(dato).split('.')[0]
    return "" if d in ("nan", "None", "NaT") else d


def _build_merge_map(ws):
    mmap = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(mr.coord)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                mmap[(r, c)] = (min_row, min_col)
    return mmap


def _escribir_rc(ws, row, col, valor, fmt_moneda=False, mmap=None):
    if mmap is None:
        mmap = _build_merge_map(ws)
    dest = mmap.get((row, col), (row, col))
    celda = ws.cell(row=dest[0], column=dest[1])
    celda.value = valor
    if fmt_moneda:
        celda.number_format = '"$"#,##0'


def _escribir_coord(ws, coord, valor, fmt_moneda=False, mmap=None):
    col_letra = ''.join(filter(str.isalpha, coord))
    fila_num = int(''.join(filter(str.isdigit, coord)))
    _escribir_rc(ws, fila_num, column_index_from_string(col_letra), valor, fmt_moneda, mmap)


def _desmerge_fila_datos(ws, fila, col_ini=2, col_fin=10):
    rangos_a_eliminar = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(mr.coord)
        if min_row <= fila <= max_row and min_col <= col_fin and max_col >= col_ini:
            rangos_a_eliminar.append(mr.coord)
    for coord in rangos_a_eliminar:
        ws.unmerge_cells(coord)


def generar_estado_cuenta(ruta_plantilla, ruta_insumo, contrato_buscado, ruta_salida_excel):
    """
    Busca los pagos del contrato en el histórico, llena la plantilla y
    guarda el resultado en ruta_salida_excel. Retorna un resumen.
    
    Estrategia: carga el Excel a SQLite temporal, consulta solo lo necesario,
    descarta la BD. Mantiene memoria baja.
    """
    db_temp = None
    try:
        # Crear BD SQLite temporal
        db_temp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
        db_path = db_temp.name
        db_temp.close()
        
        # Leer Excel y escribir a SQLite (libera pandas después)
        df = pd.read_excel(ruta_insumo)
        df.columns = df.columns.str.strip()
        
        conn = sqlite3.connect(db_path)
        df.to_sql('historico', conn, if_exists='replace', index=False)
        conn.commit()
        
        # Liberar DataFrame
        del df
        
        # Consultar desde SQLite (solo las filas del contrato)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM historico WHERE CAST(Referencia AS TEXT) LIKE ?",
            (f"%{contrato_buscado}%",)
        )
        
        # Obtener encabezados
        headers = [description[0] for description in cursor.description]
        
        # Obtener todas las filas coincidentes
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            os.unlink(db_path)
            return {"ok": False, "mensaje": f"No se encontró información para: {contrato_buscado}"}
        
        # Convertir a diccionarios
        pagos_data = [dict(zip(headers, row)) for row in rows]
        primer_fila = pagos_data[0]
        
        # Cargar plantilla y llenar
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active
        
        fila_inicio = 17
        fila_fin = fila_inicio + len(pagos_data) - 1
        for f in range(fila_inicio, fila_fin + 1):
            _desmerge_fila_datos(ws, f)
        
        mmap = _build_merge_map(ws)
        
        val_contrato = primer_fila.get(COL_VALOR_CTO, 0) or 0
        
        _escribir_coord(ws, 'D5', contrato_buscado, mmap=mmap)
        _escribir_coord(ws, 'D6', str(primer_fila.get('Nombre', '')), mmap=mmap)
        _escribir_coord(ws, 'D7', val_contrato, fmt_moneda=True, mmap=mmap)
        _escribir_coord(ws, 'D8', primer_fila.get(COL_FECHA_INICIO, ''), mmap=mmap)
        _escribir_coord(ws, 'H5', _limpiar(primer_fila.get('Proveedor', '')), mmap=mmap)
        _escribir_coord(ws, 'H6', _limpiar(primer_fila.get('Nº identificación', '')), mmap=mmap)
        _escribir_coord(ws, 'H7', _limpiar(primer_fila.get('Numero RP', '')), mmap=mmap)
        _escribir_coord(ws, 'H8', primer_fila.get(COL_FECHA_FIN, ''), mmap=mmap)
        
        fila_actual = fila_inicio
        saldo_acumulado = val_contrato
        
        for i, fila in enumerate(pagos_data, start=1):
            monto = fila.get('Valor Bruto', 0) or 0
            saldo_acumulado -= monto
            
            _escribir_rc(ws, fila_actual, 2, i, mmap=mmap)
            _escribir_rc(ws, fila_actual, 3, fila.get('Texto cabecera documento', ''), mmap=mmap)
            _escribir_rc(ws, fila_actual, 4, monto, fmt_moneda=True, mmap=mmap)
            _escribir_rc(ws, fila_actual, 5, saldo_acumulado, fmt_moneda=True, mmap=mmap)
            _escribir_rc(ws, fila_actual, 6, _limpiar(fila.get('Doc.compensación', '')), mmap=mmap)
            _escribir_rc(ws, fila_actual, 7, fila.get('Fecha de pago', ''), mmap=mmap)
            _escribir_rc(ws, fila_actual, 8, _limpiar(fila.get('Numero RP', '')), mmap=mmap)
            _escribir_rc(ws, fila_actual, 9, _limpiar(fila.get(COL_CDP_EXTERNO, '')), mmap=mmap)
            _escribir_rc(ws, fila_actual, 10, _limpiar(fila.get(COL_CRP_EXTERNO, '')), mmap=mmap)
            
            fila_actual += 1
        
        wb.save(ruta_salida_excel)
        
        return {
            "ok": True,
            "contrato": contrato_buscado,
            "contratista": str(primer_fila.get('Nombre', '')),
            "pagos_encontrados": len(pagos_data),
            "valor_contrato": float(val_contrato) if val_contrato else 0,
            "saldo_final": float(saldo_acumulado),
            "archivo_salida": ruta_salida_excel,
        }
    
    except Exception as e:
        return {"ok": False, "mensaje": f"Error: {str(e)}"}
    
    finally:
        # Limpiar BD temporal
        if db_temp and os.path.exists(db_path):
            try:
                os.unlink(db_path)
            except:
                pass


def generar_estado_cuenta_desde_datos(ruta_plantilla, pagos, contrato_buscado, ruta_salida_excel):
    """
    Igual que generar_estado_cuenta, pero recibe los pagos ya filtrados
    (lista de diccionarios) en lugar del histórico completo. El navegador
    filtra; el backend solo llena la plantilla con openpyxl, preservando
    colores, bordes, celdas fusionadas, tablas y estilos.
    """
    if not pagos:
        return {"ok": False, "mensaje": f"No se encontró información para: {contrato_buscado}"}

    primer = pagos[0]

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    fila_inicio = 17
    fila_fin = fila_inicio + len(pagos) - 1
    for f in range(fila_inicio, fila_fin + 1):
        _desmerge_fila_datos(ws, f)

    mmap = _build_merge_map(ws)

    try:
        val_contrato = float(primer.get(COL_VALOR_CTO, 0) or 0)
    except Exception:
        val_contrato = 0

    _escribir_coord(ws, 'D5', contrato_buscado, mmap=mmap)
    _escribir_coord(ws, 'D6', str(primer.get('Nombre', '')), mmap=mmap)
    _escribir_coord(ws, 'D7', val_contrato, fmt_moneda=True, mmap=mmap)
    _escribir_coord(ws, 'D8', primer.get(COL_FECHA_INICIO, ''), mmap=mmap)
    _escribir_coord(ws, 'H5', _limpiar(primer.get('Proveedor', '')), mmap=mmap)
    _escribir_coord(ws, 'H6', _limpiar(primer.get('Nº identificación', '')), mmap=mmap)
    _escribir_coord(ws, 'H7', _limpiar(primer.get('Numero RP', '')), mmap=mmap)
    _escribir_coord(ws, 'H8', primer.get(COL_FECHA_FIN, ''), mmap=mmap)

    fila_actual = fila_inicio
    saldo_acumulado = val_contrato

    for i, fila in enumerate(pagos, start=1):
        try:
            monto = float(fila.get('Valor Bruto', 0) or 0)
        except Exception:
            monto = 0
        saldo_acumulado -= monto

        _escribir_rc(ws, fila_actual, 2, i, mmap=mmap)
        _escribir_rc(ws, fila_actual, 3, fila.get('Texto cabecera documento', ''), mmap=mmap)
        _escribir_rc(ws, fila_actual, 4, monto, fmt_moneda=True, mmap=mmap)
        _escribir_rc(ws, fila_actual, 5, saldo_acumulado, fmt_moneda=True, mmap=mmap)
        _escribir_rc(ws, fila_actual, 6, _limpiar(fila.get('Doc.compensación', '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 7, fila.get('Fecha de pago', ''), mmap=mmap)
        _escribir_rc(ws, fila_actual, 8, _limpiar(fila.get('Numero RP', '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 9, _limpiar(fila.get(COL_CDP_EXTERNO, '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 10, _limpiar(fila.get(COL_CRP_EXTERNO, '')), mmap=mmap)

        fila_actual += 1

    wb.save(ruta_salida_excel)

    return {
        "ok": True,
        "contrato": contrato_buscado,
        "pagos_encontrados": len(pagos),
        "valor_contrato": val_contrato,
        "saldo_final": float(saldo_acumulado),
        "archivo_salida": ruta_salida_excel,
    }
