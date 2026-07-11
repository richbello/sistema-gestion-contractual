"""
Módulo 2 — Generación de plantilla de pagos.
Versión actualizada: usa detección dinámica de columnas, soporta múltiples
retenciones adicionales (filas código 44), y es compatible con la salida
actualizada del Módulo 01.
"""
import os
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Cruce con el registro depurado de cuentas bancarias (a prueba de fallos:
# si el servicio no estuviera disponible, buscar_cuenta devuelve None y el
# modulo sigue funcionando exactamente como antes, usando el Excel).
try:
    from .cuentas_bancarias_service import buscar_cuenta
except Exception:
    def buscar_cuenta(_cedula):
        return None

EQUIVALENCIAS = {
    "0,100%":"01","0,050%":"02","0,200%":"03","0,110%":"06",
    "2,000%":"08","0,350%":"10","0,400%":"11","1,000%":"12",
    "0,010%":"13","0,150%":"15","0,250%":"16","0,600%":"18",
    "1,100%":"22","0,700%":"26","1,104%":"33","1,380%":"34",
    "0,414%":"35","0,690%":"36","0,800%":"38","0,966%":"39",
    "1,500%":"40","0,500%":"42","0,712%":"86","0,766%":"87",
    "0,866%":"88","0,998%":"89","1,014%":"91","1,200%":"92",
    "1,214%":"R5","1,400%":"94","0,760%":"R3","0,736%":"96",
    "1,030%":"97","1,062%":"R4","1,176%":"98","1,254%":"99"
}

HEADERS = [
    'Tipo Registro P','Clave Contab.','Codigo de la cuenta','Tipo Ident',
    'No Identificación','Indicador CME','Cuenta contable','importe',
    'Indicador de IVA','RP Doc Presupuestal','Posc Doc Pres','Pros Pre',
    'Programa de financiación','Fondo','Centro Gestor','Centro de costo',
    'Centro Beneficio','Orden CO','Elemento PEP','Grafo','Area funcional',
    'Segmento','Fecha Base','Condicion de Pago','Asignación','Texto',
    'Bloqueo Pago','Receptor Alternativo','Tipo Ident','No Identificación',
    'Via de Pago','Banco Propio','Id Cta','Ref 1','Ref 2','Referencia Pago',
    'Código Bco','No Cuenta','Tipo Cta','Tipo de retenciones',
    'Indicador de retención','Base imponible de retención','Importe de retención'
]


def _float(val):
    if val is None or str(val).strip() in ('', 'nan', 'None', 'N/A'):
        return 0.0
    try:
        return float(str(val).replace(',', '.'))
    except ValueError:
        return 0.0


def _fmt_fecha(f):
    f = str(f).split(' ')[0]
    if '-' in f:
        p = f.split('-')
        if len(p) == 3:
            return f"{p[2]}/{p[1]}/{p[0]}"
    return f


def _col(df_cols, *keywords, exact=False):
    kw_up = [k.upper() for k in keywords]
    for c in df_cols:
        cu = c.upper()
        if exact:
            if cu == kw_up[0]:
                return c
        else:
            if all(k in cu for k in kw_up):
                return c
    return None


def _val(row, col_name):
    if col_name is None:
        return None
    v = row.get(col_name)
    return v if pd.notna(v) else None


def generar_plantilla_pagos(ruta_entrada_excel, ruta_salida_excel):
    df = pd.read_excel(ruta_entrada_excel, dtype=str)
    dc = list(df.columns)

    tiene_honor = 'HONOR_BASE' in dc and 'HONOR_VALOR' in dc

    C_CONTRATISTA  = _col(dc, 'CONTRATISTA')
    C_NIT          = next((c for c in dc if c.strip().upper() in ['NIT_CC','NIT','CC']), None)
    C_VALOR_BRUTO  = _col(dc, 'VALOR', 'BRUTO')
    C_BASE_RETEICA = _col(dc, 'BASE', 'RETEICA')
    C_TOTAL_DESC   = _col(dc, 'TOTAL', 'DESCUENTO')
    C_VAL_RETEICA  = _col(dc, 'VALOR', 'RETEICA')
    C_CRP          = _col(dc, 'CRP', exact=True) or _col(dc, 'CRP')
    C_CONTRATO     = _col(dc, 'CONTRATO')
    C_BCO          = _col(dc, 'CÓDIGO', 'BCO') or _col(dc, 'COD', 'BCO')
    C_CUENTA       = _col(dc, 'NO', 'CUENTA') or _col(dc, 'CUENTA')
    C_TIPO_CTA     = _col(dc, 'TIPO', 'CTA')
    C_DEL          = next((c for c in dc if c.strip().upper() == 'DEL'), None)
    C_AL           = next((c for c in dc if c.strip().upper() == 'AL'), None)
    C_PAGO_NO      = next((c for c in dc if 'PAGO' in c.upper() and 'N' in c.upper()), None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for col_num, h in enumerate(HEADERS, 1):
        ws.cell(1, col_num, h).font = Font(bold=True)

    fila_actual = 2
    filas_44    = 0
    desde_registro = 0   # cuentas tomadas del registro depurado
    sin_cruzar     = []  # (cedula, nombre) que NO estaban en el registro
    hoy         = datetime.now().strftime("%Y%m%d")

    for idx, row in df.iterrows():
        pago_num = idx + 1
        try:
            num_pago = int(float(str(_val(row, C_PAGO_NO) or pago_num)))
        except (ValueError, TypeError):
            num_pago = pago_num

        contratista  = str(_val(row, C_CONTRATISTA) or "").strip()
        no_ident     = re.sub(r'[^\d]', '',
                              str(_val(row, C_NIT) or "").split('.')[0].strip())
        valor_bruto  = _float(_val(row, C_VALOR_BRUTO))
        base_reteica = _float(_val(row, C_BASE_RETEICA))
        total_desc   = _float(_val(row, C_TOTAL_DESC))

        pct_raw = str(_val(row, C_VAL_RETEICA) or "0").replace(',', '.')
        try:
            pct_f   = float(pct_raw)
            pct_str = f"{pct_f:.3f}".replace('.', ',') + "%"
        except ValueError:
            pct_str = "0,766%"
        indicador_ret = EQUIVALENCIAS.get(pct_str, "39")

        # Retenciones adicionales → filas código 44
        rets_adicionales = []
        if tiene_honor:
            hb = _float(_val(row, 'HONOR_BASE'))
            hv = _float(_val(row, 'HONOR_VALOR'))
            if hv > 0:
                rets_adicionales.append((hb, hv))
        if 'RET2_BASE' in dc and 'RET2_VALOR' in dc:
            r2b = _float(_val(row, 'RET2_BASE'))
            r2v = _float(_val(row, 'RET2_VALOR'))
            if r2v > 0:
                rets_adicionales.append((r2b, r2v))

        rp_doc = re.sub(r'[^\d]', '',
                        str(_val(row, C_CRP) or '').split('.')[0].strip())

        contrato_raw = str(_val(row, C_CONTRATO) or "")
        nums         = re.findall(r'\d+', contrato_raw)
        asignacion   = (f"{nums[0]}-{nums[1]}" if len(nums) >= 2
                        else (nums[0] if nums else f"{pago_num:03d}-2026"))

        bco_raw  = re.sub(r'[^\d]', '', str(_val(row, C_BCO) or '051')).zfill(3)
        cta_raw  = re.sub(r'[^\d]', '',
                          re.sub(r'\.0$', '', str(_val(row, C_CUENTA) or '').strip()))
        tipo_cta = re.sub(r'[^\d]', '',
                          str(_val(row, C_TIPO_CTA) or '02')).strip().zfill(2)

        # --- Cruce con el registro depurado de cuentas bancarias ---
        # Si la cedula esta en el registro, sus datos mandan (fuente unica
        # y depurada). Si no esta, se respeta lo de la extraccion.
        cuenta_reg = buscar_cuenta(no_ident)
        if cuenta_reg:
            bco_raw  = re.sub(r'[^\d]', '', cuenta_reg['codigo_banco']).zfill(3)
            cta_raw  = re.sub(r'[^\d]', '', cuenta_reg['numero_cuenta'])
            tipo_cta = re.sub(r'[^\d]', '', cuenta_reg['tipo_cuenta']).zfill(2)
            desde_registro += 1
        else:
            sin_cruzar.append((no_ident, contratista))

        del_v   = _fmt_fecha(str(_val(row, C_DEL) or ''))
        al_v    = _fmt_fecha(str(_val(row, C_AL)  or ''))
        texto_z = f"Pago No. {num_pago} del {del_v} al {al_v}"

        # Fila C
        ws.cell(fila_actual, 1, 'C')
        ws.cell(fila_actual, 2, pago_num)
        ws.cell(fila_actual, 3, hoy)
        ws.cell(fila_actual, 4, 'KR')
        ws.cell(fila_actual, 5, '1001')
        ws.cell(fila_actual, 6, hoy)
        ws.cell(fila_actual, 8, 'COP')
        ws.cell(fila_actual, 10, asignacion)
        ws.cell(fila_actual, 11, contratista)

        # Fila P40
        ws.cell(fila_actual + 1, 1, 'P')
        ws.cell(fila_actual + 1, 2, 40)
        ws.cell(fila_actual + 1, 3, '5111809000')
        ws.cell(fila_actual + 1, 8, valor_bruto)
        ws.cell(fila_actual + 1, 9, 'WB')
        ws.cell(fila_actual + 1, 10, rp_doc)
        ws.cell(fila_actual + 1, 11, 1)
        ws.cell(fila_actual + 1, 26, texto_z)

        # Fila P31
        ws.cell(fila_actual + 2, 1, 'P')
        ws.cell(fila_actual + 2, 2, 31)
        ws.cell(fila_actual + 2, 4, 'CC')
        ws.cell(fila_actual + 2, 5, no_ident)
        ws.cell(fila_actual + 2, 7, '2401010100')
        ws.cell(fila_actual + 2, 8, valor_bruto)
        ws.cell(fila_actual + 2, 24, '0051')
        ws.cell(fila_actual + 2, 25, asignacion)
        ws.cell(fila_actual + 2, 26, texto_z)
        for col, val in zip([37, 38, 39], [bco_raw, cta_raw, tipo_cta]):
            c = ws.cell(fila_actual + 2, col, val)
            c.number_format = '@'
        ws.cell(fila_actual + 2, 40, indicador_ret)
        ws.cell(fila_actual + 2, 41, indicador_ret)
        ws.cell(fila_actual + 2, 42, base_reteica)
        ws.cell(fila_actual + 2, 43, total_desc)

        fila_actual += 3

        # Filas extra código 44
        for base_h, valor_h in rets_adicionales:
            ws.cell(fila_actual, 40, 44)
            ws.cell(fila_actual, 41, 44)
            ws.cell(fila_actual, 42, base_h)
            ws.cell(fila_actual, 43, valor_h)
            fila_actual += 1
            filas_44    += 1

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    wb.save(ruta_salida_excel)

    return {
        "ok"                    : True,
        "registros_procesados"  : len(df),
        "filas_totales"         : fila_actual - 2,
        "filas_con_codigo_44"   : filas_44,
        "tiene_columnas_honorarios": tiene_honor,
        "archivo_salida"        : ruta_salida_excel,
        "cuentas_desde_registro": desde_registro,
        "cuentas_sin_cruzar"    : [{"cedula": c, "nombre": n} for c, n in sin_cruzar],
    }
