"""
Extensión de Módulo 4 — Detección y generación automática de cesiones.

Analiza el histórico de pagos para identificar:
- Cedente (primer contratista con pagos)
- Cesionarios (contratistas posteriores)
- Llena automáticamente la sección CESIÓN en el Excel

Requiere que los pagos tengan campos: 'Nombre', 'Nº identificación', 'Fecha de pago'
"""

from openpyxl.utils import range_boundaries, column_index_from_string
from collections import defaultdict
import openpyxl


def _build_merge_map(ws):
    """Construye mapa de celdas fusionadas para preservar formato."""
    mmap = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(mr.coord)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                mmap[(r, c)] = (min_row, min_col)
    return mmap


def _escribir_rc(ws, row, col, valor, fmt_moneda=False, mmap=None):
    """Escribe valor en celda (respetando fusiones)."""
    if mmap is None:
        mmap = _build_merge_map(ws)
    dest = mmap.get((row, col), (row, col))
    celda = ws.cell(row=dest[0], column=dest[1])
    celda.value = valor
    if fmt_moneda:
        celda.number_format = '"$"#,##0'


def _escribir_coord(ws, coord, valor, fmt_moneda=False, mmap=None):
    """Escribe valor por coordenada (ej: D5)."""
    col_letra = ''.join(filter(str.isalpha, coord))
    fila_num = int(''.join(filter(str.isdigit, coord)))
    _escribir_rc(ws, fila_num, column_index_from_string(col_letra), valor, fmt_moneda, mmap)


def _limpiar(dato):
    """Limpia valores nulo/NaN."""
    d = str(dato).split('.')[0]
    return "" if d in ("nan", "None", "NaT") else d


def _desmerge_fila_datos(ws, fila, col_ini=2, col_fin=10):
    """Desmueve celdas en una fila para poder escribir datos."""
    rangos_a_eliminar = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(mr.coord)
        if min_row <= fila <= max_row and min_col <= col_fin and max_col >= col_ini:
            rangos_a_eliminar.append(mr.coord)
    for coord in rangos_a_eliminar:
        ws.unmerge_cells(coord)


def detectar_cesiones(pagos):
    """
    Analiza el histórico de pagos para detectar cesiones.
    
    Retorna:
    {
        'cedente': {
            'nombre': str,
            'cedula': str,
            'pagos': [list de pagos],
            'fecha_inicio': fecha,
            'fecha_fin': fecha,
            'total': float
        },
        'cesionarios': [
            {
                'nombre': str,
                'cedula': str,
                'pagos': [list de pagos],
                'fecha_inicio': fecha,
                'fecha_fin': fecha,
                'total': float
            }
        ]
    }
    """
    
    if not pagos:
        return None
    
    # Agrupar pagos por nombre de contratista (en orden de aparición)
    pagos_por_persona = defaultdict(list)
    orden_personas = []
    
    for pago in pagos:
        nombre = str(pago.get('Nombre', '')).strip()
        cedula = _limpiar(pago.get('Nº identificación', ''))
        
        if nombre and nombre not in orden_personas:
            orden_personas.append(nombre)
        
        if nombre:
            pagos_por_persona[nombre].append(pago)
    
    # Si solo hay 1 persona, no hay cesión
    if len(orden_personas) <= 1:
        return None
    
    # Primer contratista = cedente
    cedente_nombre = orden_personas[0]
    cedente_pagos = pagos_por_persona[cedente_nombre]
    
    resultado = {
        'cedente': {
            'nombre': cedente_nombre,
            'cedula': _limpiar(cedente_pagos[0].get('Nº identificación', '')),
            'pagos': cedente_pagos,
            'fecha_inicio': cedente_pagos[0].get('Fecha de pago', ''),
            'fecha_fin': cedente_pagos[-1].get('Fecha de pago', ''),
            'total': sum(float(p.get('Valor Bruto', 0) or 0) for p in cedente_pagos)
        },
        'cesionarios': []
    }
    
    # Resto = cesionarios
    for nombre in orden_personas[1:]:
        pagos_cesionario = pagos_por_persona[nombre]
        resultado['cesionarios'].append({
            'nombre': nombre,
            'cedula': _limpiar(pagos_cesionario[0].get('Nº identificación', '')),
            'pagos': pagos_cesionario,
            'fecha_inicio': pagos_cesionario[0].get('Fecha de pago', ''),
            'fecha_fin': pagos_cesionario[-1].get('Fecha de pago', ''),
            'total': sum(float(p.get('Valor Bruto', 0) or 0) for p in pagos_cesionario)
        })
    
    return resultado


def _copiar_rango_celdas(ws_source, ws_dest, fila_src, fila_dst, col_ini=2, col_fin=10):
    """Copia formato y estructura de una fila a otra."""
    for c in range(col_ini, col_fin + 1):
        src_cell = ws_source.cell(row=fila_src, column=c)
        dst_cell = ws_dest.cell(row=fila_dst, column=c)
        
        if src_cell.has_style:
            dst_cell.font = src_cell.font.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = src_cell.protection.copy()
            dst_cell.alignment = src_cell.alignment.copy()


def _insertar_filas_para_cesiones(ws, num_cesiones):
    """
    Inserta filas en el Excel para acomodar N cesiones.
    Cada cesión ocupa ~20 filas (datos + relación de pagos).
    """
    filas_por_cesion = 20
    filas_a_insertar = filas_por_cesion * num_cesiones
    
    # Insertar después de la fila 46 (fin de la primera cesión plantilla)
    ws.insert_rows(47, filas_a_insertar)


def llenar_seccion_cesion(ws, cesion_info, fila_inicio_datos=29, fila_inicio_pagos=40, 
                          cedente_nombre=None, cedente_valor=None, cedente_adicion=None):
    """
    Llena una sección CESIÓN completa en el Excel.
    
    Parámetros:
    - cesion_info: dict con datos del cesionario
    - fila_inicio_datos: fila donde comienzan datos básicos (29 por defecto)
    - fila_inicio_pagos: fila donde comienzan los pagos (40 por defecto)
    - cedente_nombre: nombre del cedente (para llenar "DATOS BÁSICOS")
    - cedente_valor: valor del saldo disponible para cesión
    """
    
    mmap = _build_merge_map(ws)
    
    # Desmerge de las filas de relación de pagos
    for f in range(fila_inicio_pagos, fila_inicio_pagos + 10):
        _desmerge_fila_datos(ws, f)
    
    # === SECCIÓN DATOS BÁSICOS ===
    
    # Línea: CEDENTE
    _escribir_rc(ws, fila_inicio_datos, 12, "CEDENTE", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos, 13, cedente_nombre or "", mmap=mmap)
    
    # Línea: CESIONARIO
    _escribir_rc(ws, fila_inicio_datos + 1, 12, "CESIONARIO", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 1, 13, cesion_info.get('nombre', ''), mmap=mmap)
    
    # Línea: CTO Y VIG + BP SAP
    _escribir_rc(ws, fila_inicio_datos + 2, 3, "403-2025", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 2, 12, "RP SAP", mmap=mmap)
    
    # Línea: CONTRATISTA + CC/NIT
    _escribir_rc(ws, fila_inicio_datos + 3, 3, "CONTRATISTA", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 3, 4, cesion_info.get('nombre', ''), mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 3, 7, "CC/NIT", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 3, 8, cesion_info.get('cedula', ''), mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 3, 12, "SALDO RP", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 3, 13, 0, fmt_moneda=True, mmap=mmap)
    
    # Línea: VALOR CESIÓN + RP SAP
    total_cesion = cesion_info.get('total', 0)
    _escribir_rc(ws, fila_inicio_datos + 4, 3, "VALOR CESION", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 4, 4, total_cesion, fmt_moneda=True, mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 4, 7, "RP SAP", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 4, 12, "VALOR PENDIENTE A CEDENTE", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 4, 13, 0, fmt_moneda=True, mmap=mmap)
    
    # Línea: FECHA INICIO / FECHA FINAL + SALDO PARA CESIONARIO
    _escribir_rc(ws, fila_inicio_datos + 5, 3, "FECHA INICIO", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 5, 4, cesion_info.get('fecha_inicio', ''), mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 5, 7, "FECHA FINAL", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 5, 8, cesion_info.get('fecha_fin', ''), mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 5, 12, "SALDO PARA CESIONARIO", mmap=mmap)
    _escribir_rc(ws, fila_inicio_datos + 5, 13, total_cesion, fmt_moneda=True, mmap=mmap)
    
    # === SECCIÓN RELACIÓN DE PAGOS ===
    
    fila_actual = fila_inicio_pagos
    saldo_acumulado = total_cesion
    
    for i, pago in enumerate(cesion_info.get('pagos', []), start=1):
        try:
            monto = float(pago.get('Valor Bruto', 0) or 0)
        except:
            monto = 0
        
        saldo_acumulado -= monto
        
        _escribir_rc(ws, fila_actual, 2, i, mmap=mmap)
        _escribir_rc(ws, fila_actual, 3, pago.get('Texto cabecera documento', ''), mmap=mmap)
        _escribir_rc(ws, fila_actual, 4, monto, fmt_moneda=True, mmap=mmap)
        _escribir_rc(ws, fila_actual, 5, saldo_acumulado, fmt_moneda=True, mmap=mmap)
        _escribir_rc(ws, fila_actual, 6, _limpiar(pago.get('Doc.compensación', '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 7, pago.get('Fecha de pago', ''), mmap=mmap)
        _escribir_rc(ws, fila_actual, 8, _limpiar(pago.get('Numero RP', '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 9, _limpiar(pago.get('CDP Externo', '')), mmap=mmap)
        _escribir_rc(ws, fila_actual, 10, _limpiar(pago.get('CRP Externo', '')), mmap=mmap)
        
        fila_actual += 1


def generar_estado_cuenta_con_cesiones(ruta_plantilla, pagos, contrato_buscado, 
                                        ruta_salida_excel, cedente_valor=None):
    """
    Genera estado de cuenta detectando y llenando automáticamente cesiones.
    
    DIFERENCIA con generar_estado_cuenta_desde_datos:
    - Detecta si hay múltiples contratistas (=cesiones)
    - Llena automáticamente la sección CESIÓN
    - Duplica el cuadro CESIÓN si hay N>1 cesionarios
    
    Parámetros:
    - pagos: lista de dicts con los pagos (del navegador, ya filtrados)
    - cedente_valor: valor disponible del cedente para cesión (si aplica)
    """
    
    if not pagos:
        return {"ok": False, "mensaje": f"No se encontró información para: {contrato_buscado}"}
    
    # Detectar cesiones
    cesiones_detectadas = detectar_cesiones(pagos)
    
    # Cargar plantilla
    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active
    
    primer = pagos[0]
    
    # === LLENAR SECCIÓN DE DATOS BÁSICOS ORIGINALES (primera parte del estado de cuenta) ===
    # Esto ya está en la plantilla, no necesita cambios
    
    # === LLENAR CESIONES ===
    
    if cesiones_detectadas:
        # Obtener nombre del cedente
        cedente_nombre = cesiones_detectadas['cedente']['nombre']
        cedente_valor = cedente_valor or cesiones_detectadas['cedente']['total']
        
        num_cesionarios = len(cesiones_detectadas['cesionarios'])
        
        # Si hay más de 1 cesionario, insertar filas adicionales
        if num_cesionarios > 1:
            _insertar_filas_para_cesiones(ws, num_cesionarios - 1)
        
        # Llenar cada cesión
        fila_datos = 29
        fila_pagos = 40
        
        for idx, cesionario in enumerate(cesiones_detectadas['cesionarios']):
            # Si no es la primera cesión, ajustar filas para acomodar la anterior
            if idx > 0:
                fila_datos += 20
                fila_pagos += 20
            
            llenar_seccion_cesion(
                ws,
                cesionario,
                fila_inicio_datos=fila_datos,
                fila_inicio_pagos=fila_pagos,
                cedente_nombre=cedente_nombre,
                cedente_valor=cedente_valor
            )
    
    # Guardar archivo
    wb.save(ruta_salida_excel)
    
    resultado = {
        "ok": True,
        "contrato": contrato_buscado,
        "pagos_encontrados": len(pagos),
        "cesiones_detectadas": bool(cesiones_detectadas),
        "archivo_salida": ruta_salida_excel,
    }
    
    if cesiones_detectadas:
        resultado['cedente'] = cesiones_detectadas['cedente']['nombre']
        resultado['cesionarios'] = [c['nombre'] for c in cesiones_detectadas['cesionarios']]
    
    return resultado