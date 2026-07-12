# -*- coding: utf-8 -*-
"""
Controlador (Blueprint): Validacion de plantilla de pagos.
Se monta bajo /api/validacion-plantilla
V1: caracteres/SAP + estructura.  V2: BP creado.  V3: recursos CRP.
"""
import os
import re
import traceback
from flask import Blueprint, request, jsonify, current_app
from openpyxl import load_workbook
from ..services.validacion_plantilla_service import validar_plantilla
from ..services.validacion_bp_crp_service import validar_bp_crp
from .utils import guardar_archivo

validacion_bp = Blueprint("validacion_plantilla", __name__)


def _digitos(v):
    return re.sub(r"\D", "", str(v or ""))


def _norm(v):
    return re.sub(r"\s+", " ", str(v or "").strip().upper())


def _extraer_pagos(ruta):
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(min_row=2))
    pagos = []
    i, n = 0, len(filas)
    while i < n:
        r = filas[i]
        if str(r[0].value or "").strip().upper() == "C":
            r40 = filas[i + 1] if i + 1 < n else None
            r31 = filas[i + 2] if i + 2 < n else None
            pagos.append({
                "contrato": _norm(r[9].value),
                "cedula": _digitos(r31[4].value) if r31 else "",
                "crp": _digitos(r40[9].value) if r40 else "",
                "importe": (r31[7].value if r31 else 0) or 0,
                "contratista": r[10].value or "",
            })
            i += 3
        else:
            i += 1
    return pagos


@validacion_bp.route("/procesar", methods=["POST"])
def procesar():
    try:
        upload_dir = current_app.config["UPLOAD_FOLDER"]
        if "plantilla" not in request.files:
            return jsonify({"ok": False, "mensaje": "Falta el archivo de plantilla."}), 400
        ruta = guardar_archivo(request.files["plantilla"], upload_dir)

        resultado = validar_plantilla(ruta)

        try:
            pagos = _extraer_pagos(ruta)
            bp_crp = validar_bp_crp(pagos)
            resultado["bp"] = bp_crp["bp"]
            resultado["crp"] = bp_crp["crp"]
            resultado["resumen"]["bp_sin_crear"] = sum(1 for x in bp_crp["bp"] if x["estado"] != "OK")
            resultado["resumen"]["crp_con_problema"] = sum(1 for x in bp_crp["crp"] if x["estado"] != "OK")
        except Exception as e:
            current_app.logger.error("Error en V2/V3: " + traceback.format_exc())
            resultado["bp"] = []
            resultado["crp"] = []
            resultado["aviso_bp_crp"] = f"No se pudieron ejecutar las validaciones BP/CRP: {e}"

        try:
            os.remove(ruta)
        except OSError:
            pass
        return jsonify(resultado)
    except Exception as e:
        current_app.logger.error(traceback.format_exc())
        return jsonify({"ok": False, "mensaje": f"Error validando: {e}"}), 500
