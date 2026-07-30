"""
Modulo 04 - Estado de Cuenta
============================
Version SIN PANDAS. Lectura streaming con openpyxl read_only + values_only.

Pico de memoria medido con archivos reales (historico 16 MB / 35.441 filas x 80 cols,
CRP 3.8 MB, SECOP2 326 KB):  ~110 MB
La version anterior con pandas+calamine consumia 442 MB (OOM en Render free 512 MB).
`usecols` NO reduce la memoria con calamine: parsea toda la hoja y descarta despues.

Reglas de negocio (verificadas contra el contrato 403-2025):
  - CTO Y VIG cruza contra 'No. Compromiso' por prefijo (403-2025, 403-20251, ...).
  - Se excluyen reservas (REEMPLAZA / OBLIGACION POR PAGAR / CONSTITUIRSE).
  - Adiciones se consolidan por N Interno CRP.
  - Cesion: doble confirmacion -> BP distinto en la adicion del CRP
    Y 2+ BP distintos en los pagos del historico. El cruce es por BP, no por nombre
    (el historico trunca los nombres).
  - VALOR CESION = valor inicial CRP - suma de pagos del cedente.
  - VALOR ADICION = valor de la fila 'Adicion y Prorroga' del cesionario.
  - Bloque principal: FECHA FINAL = FECHA DE TERMINACION INICIAL (antes de prorroga).
  - Bloque cesion:    FECHA FINAL = FECHA DE TERMINACION FINAL.
"""

from flask import Blueprint, request, jsonify, send_file
from flask_cors import cross_origin
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from datetime import datetime, date, timedelta
import unicodedata, copy, re, os, tempfile, json, traceback

estado_cuenta_bp = Blueprint('estado_cuenta', __name__)

# ============================ CONFIG ============================
CRP_COMPROMISO   = 'No. Compromiso'
CRP_OBJETO       = 'Objeto'
CRP_VALOR        = 'Valor CRP'
CRP_INTERNO      = 'N° Interno CRP'
CRP_BENEF_NOMBRE = 'Nombre BP Beneficiario'
CRP_BENEF_DOC    = 'Número Doc. BP Beneficiario'
CRP_BENEF_BP     = 'BP Beneficiario'

CON_CONTRATO       = 'CONTRATO'
CON_NOMBRE         = 'NOMBRE DEL CONTRATISTA'
CON_DOC            = 'NUMERO DE IDENTIFICACION'
CON_VALOR_INI      = 'VALOR INICIAL DEL CONTRATO'
CON_FECHA_INI      = 'FECHA INICIAL DE CONTRATO'
CON_FECHA_TERM_INI = 'FECHA DE TERMINACION INICIAL'
CON_FECHA_FIN      = 'FECHA DE TERMINACION FINAL'

HIS_REFERENCIA = 'Referencia'
HIS_VALOR      = 'Valor Bruto'
HIS_PERIODO    = 'Texto cabecera documento'
HIS_DOC        = 'Doc.compensación'
HIS_FECHA      = 'Fecha de pago'
HIS_PROVEEDOR  = 'Proveedor'
HIS_RP         = 'Numero RP'
HIS_CDP        = 'CDP Externo'
HIS_CRP        = 'CRP Externo'
HIS_NOMBRE     = 'Nombre'
HIS_STATUS     = 'Estatus'
HIS_STATUS_OK  = 'PAGADA'

CRP_COLS = {CRP_COMPROMISO, CRP_OBJETO, CRP_VALOR, CRP_INTERNO,
            CRP_BENEF_NOMBRE, CRP_BENEF_DOC, CRP_BENEF_BP}
HIS_COLS = {HIS_REFERENCIA, HIS_VALOR, HIS_PERIODO, HIS_DOC, HIS_FECHA,
            HIS_PROVEEDOR, HIS_RP, HIS_CDP, HIS_CRP, HIS_NOMBRE, HIS_STATUS}
CON_COLS = {CON_CONTRATO, CON_NOMBRE, CON_DOC, CON_VALOR_INI,
            CON_FECHA_INI, CON_FECHA_TERM_INI, CON_FECHA_FIN}

C_CTO, C_CONTRATISTA, C_CCNIT = 'D5', 'D6', 'H6'
C_BPSAP                       = 'H5'
C_VALOR, C_RPSAP1             = 'D7', 'H7'
C_FECHA_INI, C_FECHA_FIN      = 'D8', 'H8'
FILA_ADICION_1       = 9
FILA_PAGOS_INI_BASE  = 17
SLOTS_PAGOS_BASE     = 8      # filas de pago que trae el formato en blanco
FMT_MONEDA           = '"$"#,##0'
REEMPLAZO = re.compile(r'REEMPLAZA|OBLIGACION POR PAGAR|CONSTITUIRSE')
# ===============================================================


# --------------------------- utilidades ---------------------------
def _sinac(s):
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s))
                   if not unicodedata.combining(c)).upper()

def _es_match(nc, contrato):
    nc = str(nc).strip()
    return nc == contrato or (nc.startswith(contrato) and nc[len(contrato):].isdigit())

def _doc(v):
    """Numero sin el .0 que deja la lectura como float."""
    if v is None or v == '':
        return ''
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v).strip()

def _fecha(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    # ISO con T (2025-04-22T00:00:00.000Z) -> quedarse con la fecha
    if 'T' in s:
        s = s.split('T')[0]
    else:
        s = s.split(' ')[0]
    for f in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, f)
        except ValueError:
            pass
    return None


# ------------------- lectura streaming (sin pandas) -------------------
def _leer_filtrado(path, col_filtro, contrato, cols, match_exacto=False):
    """Lee la hoja fila por fila y devuelve solo las que casan con el contrato.

    read_only=True + values_only=True mantiene la memoria plana: openpyxl no
    materializa la hoja ni crea objetos Cell.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            hdr = next(it)
        except StopIteration:
            return []
        idx = {}
        for i, v in enumerate(hdr):
            if v is None:
                continue
            n = str(v).strip()
            if n in cols:
                idx[n] = i
        cf = idx.get(col_filtro)
        if cf is None:
            raise ValueError(f"No se encontró la columna '{col_filtro}' en el archivo")
        out = []
        for row in it:
            if cf >= len(row):
                continue
            v = row[cf]
            if v is None:
                continue
            s = str(v).strip()
            ok = _es_match(s, contrato) if match_exacto else (contrato in s)
            if ok:
                out.append({n: (row[i] if i < len(row) else None) for n, i in idx.items()})
        return out
    finally:
        wb.close()


def leer_crp(path, contrato):
    return _leer_filtrado(path, CRP_COMPROMISO, contrato, CRP_COLS, match_exacto=True)


def leer_pagos(path, contrato):
    filas = _leer_filtrado(path, HIS_REFERENCIA, contrato, HIS_COLS)
    pagos = []
    for f in filas:
        if str(f.get(HIS_STATUS) or '').strip().upper() != HIS_STATUS_OK:
            continue
        pagos.append({
            'periodo': str(f.get(HIS_PERIODO) or ''),
            'valor':   float(f.get(HIS_VALOR) or 0),
            'doc':     _doc(f.get(HIS_DOC)),
            'fecha':   _fecha(f.get(HIS_FECHA)),
            'rp':      _doc(f.get(HIS_RP)),
            'cdp':     _doc(f.get(HIS_CDP)),
            'crp':     _doc(f.get(HIS_CRP)),
            'bp':      _doc(f.get(HIS_PROVEEDOR)),
            'nombre':  str(f.get(HIS_NOMBRE) or ''),
        })
    # el historico llega desordenado: ordenar por fecha de pago
    pagos.sort(key=lambda p: p['fecha'] or datetime.max)
    return pagos


def leer_secop(path, contrato):
    filas = _leer_filtrado(path, CON_CONTRATO, contrato, CON_COLS, match_exacto=True)
    if not filas:
        return None
    r = filas[0]
    return {
        'nombre':     str(r.get(CON_NOMBRE) or ''),
        'doc':        _doc(r.get(CON_DOC)),
        'valor_ini':  float(r.get(CON_VALOR_INI) or 0),
        'f_ini':      _fecha(r.get(CON_FECHA_INI)),
        'f_term_ini': _fecha(r.get(CON_FECHA_TERM_INI)),
        'f_fin':      _fecha(r.get(CON_FECHA_FIN)),
    }


# ------------------------ logica de negocio ------------------------
def partir_crp(filas_crp):
    """Separa base de adiciones y descarta reservas."""
    base, adic = [], []
    for f in filas_crp:
        o = _sinac(f.get(CRP_OBJETO))
        if REEMPLAZO.search(o):
            continue
        (adic if 'ADICION Y PRORROGA' in o else base).append(f)
    return base, adic


def resumen_crp(base, adic):
    valor_ini = sum(float(f.get(CRP_VALOR) or 0) for f in base)
    internos = {}
    for f in base:
        k = _doc(f.get(CRP_INTERNO))
        internos[k] = internos.get(k, 0) + float(f.get(CRP_VALOR) or 0)
    rp_sap1 = max(internos, key=internos.get) if internos else ''
    ad = {}
    for f in adic:
        k = _doc(f.get(CRP_INTERNO))
        ad[k] = ad.get(k, 0) + float(f.get(CRP_VALOR) or 0)
    return {
        'valor_ini': valor_ini,
        'rp_sap1':   rp_sap1,
        'adiciones': [{'interno': k, 'valor': v} for k, v in sorted(ad.items())],
        'nombre':    str(base[0].get(CRP_BENEF_NOMBRE) or '') if base else '',
        'doc':       _doc(base[0].get(CRP_BENEF_DOC)) if base else '',
    }


def detectar_cesion(base, adic, pagos, valor_ini):
    """Doble confirmacion: BP distinto en la adicion del CRP + 2 o mas BP en los pagos."""
    if not base:
        return None
    bp_base = _doc(base[0].get(CRP_BENEF_BP))

    cesionarios = []
    for f in adic:
        bp = _doc(f.get(CRP_BENEF_BP))
        if bp and bp != bp_base:
            cesionarios.append({
                'bp':            bp,
                'nombre':        str(f.get(CRP_BENEF_NOMBRE) or ''),
                'doc':           _doc(f.get(CRP_BENEF_DOC)),
                'valor_adicion': float(f.get(CRP_VALOR) or 0),
            })
    if not cesionarios:
        return None

    por_bp = {}
    for p in pagos:
        if p['bp']:
            por_bp.setdefault(p['bp'], []).append(p)
    if len(por_bp) < 2:
        return None

    pagos_cedente = por_bp.get(bp_base, [])
    valor_cesion = valor_ini - sum(p['valor'] for p in pagos_cedente)
    for c in cesionarios:
        c['pagos'] = por_bp.get(c['bp'], [])
        c['valor_cesion'] = valor_cesion

    return {
        'bp_cedente':     bp_base,
        'nombre_cedente': str(base[0].get(CRP_BENEF_NOMBRE) or ''),
        'doc_cedente':    _doc(base[0].get(CRP_BENEF_DOC)),
        'pagos_cedente':  pagos_cedente,
        'valor_cesion':   valor_cesion,
        'cesionarios':    cesionarios,
    }


# ----------------------- manipulacion del Excel -----------------------
def _copiar_estilo(o, d):
    if o.has_style:
        d.font = copy.copy(o.font)
        d.fill = copy.copy(o.fill)
        d.border = copy.copy(o.border)
        d.alignment = copy.copy(o.alignment)
        d.number_format = o.number_format
        d.protection = copy.copy(o.protection)

def _bump(f, R, k):
    return re.sub(r'(\$?[A-Z]{1,3}\$?)(\d+)',
                  lambda m: f"{m.group(1)}{int(m.group(2)) + k}"
                  if int(m.group(2)) >= R else m.group(0), f)

def insertar_filas(ws, R, k):
    if k <= 0:
        return
    max_row, max_col = ws.max_row, ws.max_column
    readd = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= R:
            readd.append((mr.min_row + k, mr.min_col, mr.max_row + k, mr.max_col))
            ws.unmerge_cells(str(mr))
        elif mr.max_row >= R:
            readd.append((mr.min_row, mr.min_col, mr.max_row + k, mr.max_col))
            ws.unmerge_cells(str(mr))
    for row in range(max_row, R - 1, -1):
        for col in range(1, max_col + 1):
            s = ws.cell(row=row, column=col)
            d = ws.cell(row=row + k, column=col)
            v = s.value
            if isinstance(v, str) and v.startswith('='):
                v = _bump(v, R, k)
            d.value = v
            _copiar_estilo(s, d)
            s.value = None
    for (r1, c1, r2, c2) in readd:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for row in range(max_row, R - 1, -1):
        if row in ws.row_dimensions:
            ws.row_dimensions[row + k].height = ws.row_dimensions[row].height

def eliminar_filas(ws, R, k):
    if k <= 0:
        return
    max_row, max_col = ws.max_row, ws.max_column
    readd = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= R + k:
            readd.append((mr.min_row - k, mr.min_col, mr.max_row - k, mr.max_col))
            ws.unmerge_cells(str(mr))
        elif mr.min_row >= R:
            ws.unmerge_cells(str(mr))
        elif mr.max_row >= R:
            readd.append((mr.min_row, mr.min_col, max(mr.min_row, mr.max_row - k), mr.max_col))
            ws.unmerge_cells(str(mr))

    def down(f):
        return re.sub(r'(\$?[A-Z]{1,3}\$?)(\d+)',
                      lambda m: f"{m.group(1)}{int(m.group(2)) - k}"
                      if int(m.group(2)) >= R + k else m.group(0), f)
    for row in range(R + k, max_row + 1):
        for col in range(1, max_col + 1):
            s = ws.cell(row=row, column=col)
            d = ws.cell(row=row - k, column=col)
            v = s.value
            if isinstance(v, str) and v.startswith('='):
                v = down(v)
            d.value = v
            _copiar_estilo(s, d)
    for row in range(max_row - k + 1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None
    for (r1, c1, r2, c2) in readd:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _merge_map(ws):
    """(fila, col) -> celda ancla, para escribir dentro de rangos combinados."""
    m = {}
    for mr in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(mr.coord)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                m[(r, c)] = (r1, c1)
    return m

def _wrc(ws, row, col, val, m, fmt=None):
    r, c = m.get((row, col), (row, col))
    cel = ws.cell(row=r, column=c)
    cel.value = val
    if fmt:
        cel.number_format = fmt

def _desmerge(ws, fila, c1=2, c2=10):
    for mr in list(ws.merged_cells.ranges):
        mc1, mr1, mc2, mr2 = range_boundaries(mr.coord)
        if mr1 <= fila <= mr2 and mc1 <= c2 and mc2 >= c1:
            ws.unmerge_cells(mr.coord)

def _buscar(ws, texto, col=3, desde=1, hasta=None):
    """Localiza un rotulo por texto: robusto ante desplazamiento por adiciones."""
    t = texto.strip().upper()
    hasta = hasta or ws.max_row
    for r in range(desde, hasta + 1):
        v = ws.cell(row=r, column=col).value
        if v and t in str(v).strip().upper():
            return r
    return None


def llenar_cesion(ws, ces, f_ini_ces, f_fin_ces, fila_desde):
    """Llena el bloque CESION. Devuelve True si escribio algo."""
    if not ces or not ces['cesionarios']:
        return False
    c1 = ces['cesionarios'][0]

    f_ces = _buscar(ws, 'CESIÓN', col=2, desde=fila_desde) or \
            _buscar(ws, 'CESION', col=2, desde=fila_desde)
    if not f_ces:
        return False
    f_cto = _buscar(ws, 'CTO Y VIG', col=3, desde=f_ces)
    if not f_cto:
        return False

    m = _merge_map(ws)
    _wrc(ws, f_cto,     4, ces.get('contrato', ''), m)
    _wrc(ws, f_cto,     8, c1['bp'], m, 'General')
    _wrc(ws, f_cto + 1, 4, c1['nombre'], m)
    _wrc(ws, f_cto + 1, 8, c1['doc'], m, 'General')
    _wrc(ws, f_cto + 2, 4, c1['valor_cesion'], m, FMT_MONEDA)
    _wrc(ws, f_cto + 2, 8, c1['pagos'][0]['rp'] if c1['pagos'] else '', m, 'General')
    if f_ini_ces:
        _wrc(ws, f_cto + 3, 4, f_ini_ces, m, 'DD/MM/YYYY')
    if f_fin_ces:
        _wrc(ws, f_cto + 3, 8, f_fin_ces, m, 'DD/MM/YYYY')
    _wrc(ws, f_cto + 4, 4, c1['valor_adicion'], m, FMT_MONEDA)

    f_hdr = _buscar(ws, 'PERIODO', col=3, desde=f_cto + 4)
    if not f_hdr:
        return False
    f_pago = f_hdr + 1

    f_total = _buscar(ws, 'TOTAL', col=2, desde=f_pago) or (f_pago + 5)
    slots = f_total - f_pago
    n = len(c1['pagos'])
    if n > slots:
        insertar_filas(ws, f_pago + slots, n - slots)
    elif 0 < n < slots:
        eliminar_filas(ws, f_pago + n, slots - n)
    f_total = f_pago + max(n, slots if n == 0 else n)

    m = _merge_map(ws)
    for i in range(max(n, 1)):
        _desmerge(ws, f_pago + i)
    m = _merge_map(ws)

    fmt = ws.cell(row=f_pago, column=4).number_format or FMT_MONEDA
    for i, p in enumerate(c1['pagos']):
        r = f_pago + i
        _wrc(ws, r,  2, i + 1, m)
        _wrc(ws, r,  3, p['periodo'], m)
        _wrc(ws, r,  4, p['valor'], m, fmt)
        if i == 0:
            ws.cell(row=r, column=5).value = f"=(D{f_cto + 2}+D{f_cto + 4})-D{r}"
        else:
            ws.cell(row=r, column=5).value = f"=E{r - 1}-D{r}"
        ws.cell(row=r, column=5).number_format = fmt
        _wrc(ws, r,  6, p['doc'], m, 'General')
        if p['fecha']:
            _wrc(ws, r, 7, p['fecha'], m, 'DD/MM/YYYY')
        _wrc(ws, r,  8, p['rp'], m, 'General')
        _wrc(ws, r,  9, p['cdp'], m, 'General')
        _wrc(ws, r, 10, p['crp'], m, 'General')

    if n:
        ult = f_pago + n - 1
        ws.cell(row=f_total, column=4).value = f"=SUM(D{f_pago}:D{ult})"
        ws.cell(row=f_total, column=4).number_format = fmt
        cel = ws.cell(row=f_total, column=5)
        cel.value = f"=E{ult}"
        cel.number_format = fmt
        cel.font = Font(name=cel.font.name, size=cel.font.size, bold=True)
    return True


# --------------------------- orquestacion ---------------------------
def _preparar_y_armar(plantilla_path, base, adic, pagos, sec, contrato, salida_path):
    """Toma datos YA extraidos (de archivo o de JSON) y arma el Excel. RAM minima."""
    contrato = str(contrato).strip()
    crp = resumen_crp(base, adic)

    if crp['valor_ini'] == 0 and not crp['adiciones'] and sec is None:
        return {'ok': False,
                'mensaje': f'No se encontró información para el contrato {contrato}'}

    valor_ini = crp['valor_ini'] or (sec['valor_ini'] if sec else 0.0)
    ces = detectar_cesion(base, adic, pagos, valor_ini)

    if ces:
        pagos_principal = ces['pagos_cedente']
        nombre = ces['nombre_cedente'] or (sec['nombre'] if sec else '') or crp['nombre']
        doc    = ces['doc_cedente']    or (sec['doc'] if sec else '')    or crp['doc']
        # el bloque principal cierra en la terminacion INICIAL (antes de prorroga)
        fecha_fin = (sec.get('f_term_ini') if sec else None) or (sec.get('f_fin') if sec else None)
        f_fin_ces = sec.get('f_fin') if sec else None
        f_ini_ces = (fecha_fin + timedelta(days=1)) if fecha_fin else None
    else:
        pagos_principal = pagos
        nombre = (sec['nombre'] if sec else '') or crp['nombre']
        doc    = (sec['doc'] if sec else '')    or crp['doc']
        fecha_fin = sec.get('f_fin') if sec else None
        f_ini_ces = f_fin_ces = None

    fecha_ini = sec.get('f_ini') if sec else None
    adiciones = crp['adiciones']
    n_ad = len(adiciones)
    bp_sap = pagos_principal[0]['bp'] if pagos_principal else ''

    wb = load_workbook(plantilla_path)
    ws = wb.active

    # cabecera
    ws[C_CTO] = contrato
    ws[C_CONTRATISTA] = nombre
    ws[C_CCNIT] = doc
    ws[C_CCNIT].number_format = 'General'
    if bp_sap:
        ws[C_BPSAP] = bp_sap
        ws[C_BPSAP].number_format = 'General'
    ws[C_VALOR] = valor_ini
    ws[C_VALOR].number_format = FMT_MONEDA
    if crp['rp_sap1']:
        ws[C_RPSAP1] = crp['rp_sap1']
        ws[C_RPSAP1].number_format = 'General'
    if fecha_ini:
        ws[C_FECHA_INI] = fecha_ini
        ws[C_FECHA_INI].number_format = 'DD/MM/YYYY'
    if fecha_fin:
        ws[C_FECHA_FIN] = fecha_fin
        ws[C_FECHA_FIN].number_format = 'DD/MM/YYYY'

    # adiciones
    if n_ad >= 1:
        ws[f'D{FILA_ADICION_1}'] = adiciones[0]['valor']
        ws[f'D{FILA_ADICION_1}'].number_format = FMT_MONEDA
        ws[f'H{FILA_ADICION_1}'] = adiciones[0]['interno']
        ws[f'H{FILA_ADICION_1}'].number_format = 'General'
    extra = adiciones[1:]
    if extra:
        insertar_filas(ws, FILA_ADICION_1 + 1, len(extra))
        for i, ad in enumerate(extra):
            r = FILA_ADICION_1 + 1 + i
            for cc in ('C', 'D', 'G', 'H'):
                _copiar_estilo(ws[f'{cc}{FILA_ADICION_1}'], ws[f'{cc}{r}'])
            ws[f'C{r}'] = f'VALOR ADICIÓN {i + 2}'
            ws[f'D{r}'] = ad['valor']
            ws[f'D{r}'].number_format = FMT_MONEDA
            ws[f'G{r}'] = f'RP ADICIÓN {i + 2}'
            ws[f'H{r}'] = ad['interno']
            ws[f'H{r}'].number_format = 'General'

    k_ad = len(extra)
    # localizar la tabla de pagos por rotulo (robusto ante el desplazamiento)
    f_hdr = _buscar(ws, 'PERIODO', col=3, desde=FILA_ADICION_1 + k_ad,
                    hasta=FILA_PAGOS_INI_BASE + k_ad + 10)
    pago_ini = (f_hdr + 1) if f_hdr else (FILA_PAGOS_INI_BASE + k_ad)
    f_tot = _buscar(ws, 'TOTAL', col=2, desde=pago_ini)
    slots = (f_tot - pago_ini) if f_tot else SLOTS_PAGOS_BASE

    n_pagos = len(pagos_principal)
    if n_pagos > slots:
        insertar_filas(ws, pago_ini + slots, n_pagos - slots)
    elif 0 < n_pagos < slots:
        eliminar_filas(ws, pago_ini + n_pagos, slots - n_pagos)
    filas_pago = max(n_pagos, slots if n_pagos == 0 else n_pagos)
    total_row = pago_ini + filas_pago

    base_saldo = (f"{C_VALOR}+SUM(D{FILA_ADICION_1}:D{FILA_ADICION_1 + n_ad - 1})"
                  if n_ad >= 1 else C_VALOR)
    fmt = ws.cell(row=pago_ini, column=4).number_format or FMT_MONEDA

    for i in range(filas_pago):
        r = pago_ini + i
        if i < n_pagos:
            p = pagos_principal[i]
            ws.cell(row=r, column=2).value = i + 1
            ws.cell(row=r, column=3).value = p['periodo']
            ws.cell(row=r, column=4).value = p['valor']
            ws.cell(row=r, column=4).number_format = fmt
            ws.cell(row=r, column=6).value = p['doc']
            if p['fecha']:
                ws.cell(row=r, column=7).value = p['fecha']
                ws.cell(row=r, column=7).number_format = 'DD/MM/YYYY'
            ws.cell(row=r, column=8).value = p['rp']
            ws.cell(row=r, column=9).value = p['cdp']
            ws.cell(row=r, column=10).value = p['crp']
        ws.cell(row=r, column=5).value = (f"={base_saldo}-D{r}" if i == 0
                                          else f"=E{r - 1}-D{r}")
        ws.cell(row=r, column=5).number_format = fmt

    ult = total_row - 1
    ws.cell(row=total_row, column=4).value = f"=SUM(D{pago_ini}:D{ult})"
    ws.cell(row=total_row, column=4).number_format = fmt
    cel = ws.cell(row=total_row, column=5)
    cel.value = f"=E{ult}"
    cel.number_format = fmt
    cel.font = Font(name=cel.font.name, size=cel.font.size, bold=True)

    hay_cesion = False
    if ces:
        ces['contrato'] = contrato
        hay_cesion = llenar_cesion(ws, ces, f_ini_ces, f_fin_ces, total_row)

    wb.save(salida_path)

    res = {'ok': True, 'contrato': contrato, 'contratista': nombre,
           'valor_inicial': valor_ini, 'n_adiciones': n_ad, 'n_pagos': n_pagos,
           'cesion': hay_cesion,
           'valor_final': valor_ini + sum(a['valor'] for a in adiciones)}
    if ces:
        res['cesionario'] = ces['cesionarios'][0]['nombre']
        res['valor_cesion'] = ces['valor_cesion']
        res['n_pagos_cesion'] = len(ces['cesionarios'][0]['pagos'])
    return res


def generar_estado_cuenta(plantilla_path, crp_path, consolidado_path,
                          historico_path, contrato, salida_path):
    contrato = str(contrato).strip()

    filas_crp = leer_crp(crp_path, contrato) if crp_path else []
    pagos     = leer_pagos(historico_path, contrato) if historico_path else []
    sec       = leer_secop(consolidado_path, contrato) if consolidado_path else None

    base, adic = partir_crp(filas_crp)
    return _preparar_y_armar(plantilla_path, base, adic, pagos, sec, contrato, salida_path)





# --------------------- endpoint LITE: datos ya filtrados por el navegador ---------------------
def armar_desde_datos(plantilla_path, datos, salida_path):
    """Recibe el JSON pequeno que arma el navegador (filas ya filtradas por contrato)."""
    contrato = str(datos.get('contrato', '')).strip()

    filas_crp = []
    for r in datos.get('crp', []):
        filas_crp.append({
            CRP_COMPROMISO:   r.get('compromiso', ''),
            CRP_OBJETO:       r.get('objeto', ''),
            CRP_VALOR:        r.get('valor', 0),
            CRP_INTERNO:      r.get('interno', ''),
            CRP_BENEF_NOMBRE: r.get('nombre_benef', ''),
            CRP_BENEF_DOC:    r.get('doc_benef', ''),
            CRP_BENEF_BP:     r.get('bp_benef', ''),
        })

    pagos = []
    for p in datos.get('pagos', []):
        pagos.append({
            'periodo': p.get('periodo', ''),
            'valor':   float(p.get('valor', 0) or 0),
            'doc':     _doc(p.get('doc')),
            'fecha':   _fecha(p.get('fecha')),
            'rp':      _doc(p.get('rp')),
            'cdp':     _doc(p.get('cdp')),
            'crp':     _doc(p.get('crp')),
            'bp':      _doc(p.get('bp')),
            'nombre':  p.get('nombre', ''),
        })
    pagos.sort(key=lambda p: p['fecha'] or datetime.max)

    sd = datos.get('secop')
    sec = None
    if sd:
        sec = {
            'nombre': sd.get('nombre', ''), 'doc': _doc(sd.get('doc')),
            'valor_ini': float(sd.get('valor_ini', 0) or 0),
            'f_ini': _fecha(sd.get('f_ini')),
            'f_term_ini': _fecha(sd.get('f_term_ini')),
            'f_fin': _fecha(sd.get('f_fin')),
        }

    base, adic = partir_crp(filas_crp)
    return _preparar_y_armar(plantilla_path, base, adic, pagos, sec, contrato, salida_path)


# ============================ ENDPOINTS ============================
@estado_cuenta_bp.route('/procesar', methods=['POST'])
@cross_origin()
def procesar():
    tmp = None
    try:
        plantilla   = request.files.get('plantilla')
        reporte_crp = request.files.get('reporte_crp')
        consolidado = request.files.get('consolidado')   # opcional
        historico   = request.files.get('historico')     # opcional
        contrato    = (request.form.get('contrato') or '').strip()

        if not plantilla or not reporte_crp or not contrato:
            return jsonify({'ok': False,
                            'mensaje': 'Faltan datos: se requieren plantilla, reporte_crp y contrato'}), 400

        tmp = tempfile.mkdtemp()
        p_pl = os.path.join(tmp, 'plantilla.xlsx');   plantilla.save(p_pl)
        p_crp = os.path.join(tmp, 'crp.xlsx');        reporte_crp.save(p_crp)
        p_con = p_his = None
        if consolidado:
            p_con = os.path.join(tmp, 'consolidado.xlsx'); consolidado.save(p_con)
        if historico:
            p_his = os.path.join(tmp, 'historico.xlsx');   historico.save(p_his)

        salida = os.path.join(tmp, f'Estado_de_Cuenta_{contrato}.xlsx')
        res = generar_estado_cuenta(p_pl, p_crp, p_con, p_his, contrato, salida)
        if not res.get('ok'):
            return jsonify(res), 400
        return send_file(salida, as_attachment=True,
                         download_name=f'Estado_de_Cuenta_{contrato}.xlsx')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'mensaje': str(e)}), 500


@estado_cuenta_bp.route('/procesar-lite', methods=['POST'])
def procesar_lite():
    """Ruta previa conservada (recibe pagos ya armados en JSON)."""
    try:
        from app.services.estado_cuenta_service import generar_estado_cuenta_desde_datos
    except ImportError:
        from ..services.estado_cuenta_service import generar_estado_cuenta_desde_datos
    try:
        plantilla = request.files.get('plantilla')
        contrato = (request.form.get('contrato') or '').strip()
        pagos_raw = request.form.get('pagos')
        if not plantilla or not contrato or not pagos_raw:
            return jsonify({'ok': False, 'mensaje': 'Faltan datos'}), 400
        pagos = json.loads(pagos_raw)
        tmp = tempfile.mkdtemp()
        p_pl = os.path.join(tmp, 'plantilla.xlsx')
        plantilla.save(p_pl)
        salida = os.path.join(tmp, f'Estado_de_Cuenta_{contrato}.xlsx')
        r = generar_estado_cuenta_desde_datos(p_pl, pagos, contrato, salida)
        if not r.get('ok'):
            return jsonify(r), 400
        return send_file(salida, as_attachment=True,
                         download_name=f'Estado_de_Cuenta_{contrato}.xlsx')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'mensaje': str(e)}), 500


@estado_cuenta_bp.route('/procesar-json', methods=['POST'])
@cross_origin()
def procesar_json():
    """El navegador filtra los Excel y manda solo las filas del contrato (JSON pequeno)."""
    try:
        plantilla = request.files.get('plantilla')
        datos_raw = request.form.get('datos')
        contrato  = (request.form.get('contrato') or '').strip()
        if not plantilla or not datos_raw or not contrato:
            return jsonify({'ok': False, 'mensaje': 'Faltan datos: plantilla, datos y contrato'}), 400
        datos = json.loads(datos_raw)
        datos.setdefault('contrato', contrato)
        tmp = tempfile.mkdtemp()
        p_pl = os.path.join(tmp, 'plantilla.xlsx')
        plantilla.save(p_pl)
        salida = os.path.join(tmp, f'Estado_de_Cuenta_{contrato}.xlsx')
        res = armar_desde_datos(p_pl, datos, salida)
        if not res.get('ok'):
            return jsonify(res), 400
        return send_file(salida, as_attachment=True,
                         download_name=f'Estado_de_Cuenta_{contrato}.xlsx')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'mensaje': str(e)}), 500