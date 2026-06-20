"""
Módulo 2 — Generación de plantilla de pagos consolidada.

Adaptado del script original de Colab: toma el Excel de extracción
(Módulo 1) y genera la plantilla de carga para el sistema de pagos
(formato con registros C / P40 / P31, y fila adicional código 44 cuando
hay honorarios).
"""
import os
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

EQUIVALENCIAS = {
    "0,100%": "01", "0,050%": "02", "0,200%": "03", "0,110%": "06",
    "2,000%": "08", "0,350%": "10", "0,400%": "11", "1,000%": "12",
    "0,010%": "13", "0,150%": "15", "0,250%": "16", "0,600%": "18",
    "1,100%": "22", "0,700%": "26", "1,104%": "33", "1,380%": "34",
    "0,414%": "35", "0,690%": "36", "0,800%": "38", "0,966%": "39",
    "1,500%": "40", "0,500%": "42", "0,712%": "86", "0,766%": "87",
    "0,866%": "88", "0,998%": "89", "1,014%": "91", "1,200%": "92",
    "1,214%": "R5", "1,400%": "94", "0,760%": "R3", "0,736%": "96",
    "1,030%": "97", "1,062%": "R4", "1,176%": "98", "1,254%": "99"
}

HEADERS = [
    'Tipo Registro P', 'Clave Contab.', 'Codigo de la cuenta', 'Tipo Ident',
    'No Identificación', 'Indicador CME', 'Cuenta contable', 'importe',
    'Indicador de IVA', 'RP Doc Presupuestal', 'Posc Doc Pres', 'Pros Pre',
    'Programa de financiación', 'Fondo', 'Centro Gestor', 'Centro de costo',
    'Centro Beneficio', 'Orden CO', 'Elemento PEP', 'Grafo', 'Area funcional',
    'Segmento', 'Fecha Base', 'Condicion de Pago', 'Asignación', 'Texto',
    'Bloqueo Pago', 'Receptor Alternativo', 'Tipo Ident', 'No Identificación',
    'Via de Pago', 'Banco Propio', 'Id Cta', 'Ref 1', 'Ref 2', 'Referencia Pago',
    'Código Bco', 'No Cuenta', 'Tipo Cta', 'Tipo de retenciones',
    'Indicador de retención', 'Base imponible de retención', 'Importe de retención'
]


def _fmt_fecha(f):
    f = str(f).split(' ')[0]
    if '-' in f:
        p = f.split('-')
        if len(p) == 3:
            return f"{p[2]}/{p[1]}/{p[0]}"
    return f


def generar_plantilla_pagos(ruta_entrada_excel, ruta_salida_excel):
    """
    Lee el Excel de extracción y genera la plantilla de pagos (.xlsx) en
    ruta_salida_excel. Retorna un resumen del proceso.
    """
    df = pd.read_excel(ruta_entrada_excel, dtype=str)

    tiene_col_honor = 'HONOR_BASE' in df.columns and 'HONOR_VALOR' in df.columns

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for col_num, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

    fila_actual = 2
    filas_con_44 = 0

    for idx, row in df.iterrows():
        pago_num = idx + 1
        try:
            num_pago = int(float(str(row["PAGO No."]).strip())) if pd.notna(row["PAGO No."]) else pago_num
        except (KeyError, ValueError):
            num_pago = pago_num

        contratista = next((str(row[c]).strip() for c in df.columns
                             if "CONTRATISTA" in str(c).upper() and pd.notna(row[c])), "")

        no_ident = next((str(row[c]).split('.')[0].strip() for c in df.columns
                          if str(c).strip().upper() in ['NIT_CC', 'NIT', 'CC'] and pd.notna(row[c])),
                         f"ID{pago_num:04d}")

        valor_bruto = next((str(row[c]).replace(',', '.') for c in df.columns
                             if 'valor' in str(c).lower() and 'bruto' in str(c).lower() and pd.notna(row[c])), "0")
        try:
            valor_bruto_float = float(valor_bruto)
        except ValueError:
            valor_bruto_float = 0.0

        base_ret = next((str(row[c]).replace(',', '.') for c in df.columns
                          if str(c).strip().upper() == 'BASE RETEICA' and pd.notna(row[c])), "0")
        try:
            base_ret_float = float(base_ret)
        except ValueError:
            base_ret_float = 0.0

        importe_ret = next((str(row[c]).replace(',', '.') for c in df.columns
                             if str(c).strip().upper() == 'TOTAL DESCUENTOS' and pd.notna(row[c])), "0")
        try:
            importe_ret_float = float(importe_ret)
        except ValueError:
            importe_ret_float = 0.0

        val_ret_raw = next(
            (str(row[c]) for c in df.columns if 'valor' in str(c).lower() and 'reteica' in str(c).lower() and pd.notna(row[c])),
            "0"
        )
        val_ret_pct = "0,766%"
        try:
            num = float(val_ret_raw.replace(',', '.'))
            val_ret_pct = f"{num:.3f}".replace('.', ',') + "%"
        except ValueError:
            if '%' in val_ret_raw:
                val_ret_pct = val_ret_raw.strip().replace('.', ',')

        indicador_ret = EQUIVALENCIAS.get(val_ret_pct, "39")

        honor_base_float = 0.0
        honor_valor_float = 0.0
        if tiene_col_honor:
            hb = row.get('HONOR_BASE', None)
            hv = row.get('HONOR_VALOR', None)
            if pd.notna(hb) and str(hb).strip() not in ('', 'nan'):
                try:
                    honor_base_float = float(str(hb).replace(',', '.'))
                except ValueError:
                    honor_base_float = 0.0
            if pd.notna(hv) and str(hv).strip() not in ('', 'nan'):
                try:
                    honor_valor_float = float(str(hv).replace(',', '.'))
                except ValueError:
                    honor_valor_float = 0.0

        tiene_honorarios = honor_valor_float > 0

        rp_doc = next((str(row[c]).split('.')[0].strip() for c in df.columns
                       if str(c).strip().upper() == 'CRP' and pd.notna(row[c])), "")

        contrato_raw = next((str(row[c]) for c in df.columns if 'contrato' in str(c).lower() and pd.notna(row[c])), "")
        nums = re.findall(r'\d+', contrato_raw)
        asignacion = f"{nums[0]}-{nums[1]}" if len(nums) >= 2 else (nums[0] if nums else f"{pago_num:03d}-2025")

        cod_bco_raw = next((str(row[c]) for c in df.columns
                             if 'código' in str(c).lower() and 'bco' in str(c).lower() and pd.notna(row[c])), "051")
        codigo_bco = re.sub(r'[^\d]', '', cod_bco_raw).strip().zfill(3)

        cuenta_raw = next((str(row[c]) for c in df.columns
                            if 'no' in str(c).lower() and 'cuenta' in str(c).lower() and pd.notna(row[c])), "0550488435468647")
        no_cuenta = re.sub(r'\.0$', '', cuenta_raw.strip())
        no_cuenta = re.sub(r'[^\d]', '', no_cuenta)

        tipo_cta_raw = next((str(row[c]) for c in df.columns
                              if 'tipo' in str(c).lower() and 'cta' in str(c).lower() and pd.notna(row[c])), "02")
        tipo_cta = re.sub(r'[^\d]', '', tipo_cta_raw).strip().zfill(2)

        del_v = next((str(row[c]).strip() for c in df.columns if str(c).strip().upper() == "DEL" and pd.notna(row[c])), "")
        al_v = next((str(row[c]).strip() for c in df.columns if str(c).strip().upper() == "AL" and pd.notna(row[c])), "")
        del_v = _fmt_fecha(del_v)
        al_v = _fmt_fecha(al_v)
        texto_z = f"Pago No. {num_pago} del {del_v} al {al_v}"

        # Fila C
        ws.cell(row=fila_actual, column=1, value='C')
        ws.cell(row=fila_actual, column=2, value=pago_num)
        ws.cell(row=fila_actual, column=3, value=datetime.now().strftime("%Y%m%d"))
        ws.cell(row=fila_actual, column=4, value='KR')
        ws.cell(row=fila_actual, column=5, value='1001')
        ws.cell(row=fila_actual, column=6, value=datetime.now().strftime("%Y%m%d"))
        ws.cell(row=fila_actual, column=8, value='COP')
        ws.cell(row=fila_actual, column=10, value=asignacion)
        ws.cell(row=fila_actual, column=11, value=contratista)

        # Fila P40
        ws.cell(row=fila_actual + 1, column=1, value='P')
        ws.cell(row=fila_actual + 1, column=2, value=40)
        ws.cell(row=fila_actual + 1, column=3, value='5111809000')
        ws.cell(row=fila_actual + 1, column=8, value=valor_bruto_float)
        ws.cell(row=fila_actual + 1, column=9, value='WB')
        ws.cell(row=fila_actual + 1, column=10, value=rp_doc)
        ws.cell(row=fila_actual + 1, column=11, value=1)
        ws.cell(row=fila_actual + 1, column=26, value=texto_z)

        # Fila P31
        ws.cell(row=fila_actual + 2, column=1, value='P')
        ws.cell(row=fila_actual + 2, column=2, value=31)
        ws.cell(row=fila_actual + 2, column=4, value='CC')
        ws.cell(row=fila_actual + 2, column=5, value=no_ident)
        ws.cell(row=fila_actual + 2, column=7, value='2401010100')
        ws.cell(row=fila_actual + 2, column=8, value=valor_bruto_float)
        ws.cell(row=fila_actual + 2, column=24, value='0051')
        ws.cell(row=fila_actual + 2, column=25, value=asignacion)
        ws.cell(row=fila_actual + 2, column=26, value=texto_z)

        for col, val in zip([37, 38, 39], [codigo_bco, no_cuenta, tipo_cta]):
            cell = ws.cell(row=fila_actual + 2, column=col, value=val)
            cell.number_format = '@'

        ws.cell(row=fila_actual + 2, column=40, value=indicador_ret)
        ws.cell(row=fila_actual + 2, column=41, value=indicador_ret)
        ws.cell(row=fila_actual + 2, column=42, value=base_ret_float)
        ws.cell(row=fila_actual + 2, column=43, value=importe_ret_float)

        if tiene_honorarios:
            ws.cell(row=fila_actual + 3, column=40, value=44)
            ws.cell(row=fila_actual + 3, column=41, value=44)
            ws.cell(row=fila_actual + 3, column=42, value=honor_base_float)
            ws.cell(row=fila_actual + 3, column=43, value=honor_valor_float)
            fila_actual += 4
            filas_con_44 += 1
        else:
            fila_actual += 3

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    wb.save(ruta_salida_excel)

    return {
        "ok": True,
        "registros_procesados": len(df),
        "filas_con_codigo_44": filas_con_44,
        "tiene_columnas_honorarios": tiene_col_honor,
        "archivo_salida": ruta_salida_excel,
    }
